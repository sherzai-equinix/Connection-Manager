# routers/importer.py
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from fastapi import Body
from database import get_db
from fastapi import HTTPException
from sqlalchemy import text
from security import get_current_user, require_permissions
from audit import write_audit_log

router = APIRouter(
    prefix="/import",
    tags=["import"],
)


# ============================================================
# Helpers
# ============================================================

def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()

def _lower(s: Any) -> str:
    return _norm(s).lower()

def _norm_port_label(s: Any) -> str:
    """
    Normalize port labels from Excel/user input:
      '1 B5' -> '1B5'
      ' 2c3 ' -> '2C3'
    """
    v = _norm(s).upper()
    v = re.sub(r"\s+", "", v)
    return v

def base_room(value: str) -> str:
    """
    base_room("5.13S1/RU27") -> "5.13"
    base_room("5.13") -> "5.13"
    """
    v = _norm(value)
    m = re.match(r"^(\d+\.\d+)", v)
    return m.group(1) if m else v

def _extract_cassette_letter(port_label: Any) -> str:
    """
    Extract cassette letter from a port label.
    Examples:
      '5B6' -> 'B'
      '2A2' -> 'A'
      '11C5' -> 'C'
      '1B1' -> 'B'
    """
    s = _norm(port_label).upper()
    m = re.search(r"[A-Z]", s)
    return m.group(0) if m else ""


# ============================================================
# Excel header detection + refinement
# ============================================================

def _find_header_map(row_values: List[Any]) -> Optional[Dict[str, int]]:
    """
    Find ONLY the stable headers:
      - Product ID
      - Router Port
      - HU/PP Z

    Customer + Customer Port will be refined later using sample scoring.
    """
    vals = [_lower(v) for v in row_values]

    def find_first(keys: List[str]) -> Optional[int]:
        for i, v in enumerate(vals):
            if not v:
                continue
            for k in keys:
                if k in v:
                    return i
        return None

    col_product = find_first(["product id", "productid", "product-id"])
    col_router = find_first(["router port", "routerport", "router-port"])
    col_hu_pp_z = find_first(["hu / pp z", "hu/pp z", "hu pp z", "pp z", "hu/ppz", "hu ppz"])

    col_eqx_serial = find_first(["eqx serial", "eqx", "serial"])

    if col_product is None or col_router is None or col_hu_pp_z is None:
        return None

    # placeholders for refinement
    return {
        "product_id": col_product,
        "router_port": col_router,
        "z_pp_number": col_hu_pp_z,
        "customer_name": -1,
        "z_port_label": -1,
        "eqx_serial": col_eqx_serial if col_eqx_serial is not None else -1,
    }


def _score_customer_cell(v: Any) -> int:
    s = _norm(v)
    if not s:
        return 0
    if "RFRA" in s.upper():
        return 0
    # customer names are usually text with letters
    return 3 if re.search(r"[A-Za-z]", s) else 1

def _score_zport_cell(v: Any) -> int:
    s = _norm_port_label(v)
    if not s:
        return 0
    if "RFRA" in s:
        return 0
    # like 1B1, 2C3, 6A2
    if re.match(r"^\d+[A-Z]\d+$", s):
        return 5
    return 0

def _choose_best_col(ws, header_row_idx: int, candidates: List[int], scorer, sample_rows: int = 25) -> Optional[int]:
    if not candidates:
        return None

    best_col = None
    best_score = -1
    max_r = min(ws.max_row, header_row_idx + sample_rows)

    for ci in candidates:
        score = 0
        for r in range(header_row_idx + 1, max_r + 1):
            val = ws.cell(row=r, column=ci + 1).value  # ci 0-based; ws 1-based
            score += scorer(val)
        if score > best_score:
            best_score = score
            best_col = ci

    return best_col


def parse_install_rows(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parses weekly Excel:
    - scans for header row
    - refines Customer + Customer Port columns by sampling values
    - reads rows until long empty streak
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl not installed. pip install openpyxl") from e

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header_map: Optional[Dict[str, int]] = None
    header_row_idx: Optional[int] = None

    for r in range(1, min(ws.max_row, 120) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        hm = _find_header_map(row_vals)
        if hm:
            header_map = hm
            header_row_idx = r
            break

    if not header_map or not header_row_idx:
        raise ValueError("Header nicht gefunden (Product ID / Router Port / HU/PP Z)")

    # ---- refine customer + port columns by looking at header row + samples
    header_vals = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    header_lower = [_lower(v) for v in header_vals]

    router_port_idx = int(header_map["router_port"])

    customer_candidates = [
        i for i, v in enumerate(header_lower)
        if v and ("customer" in v or "kunde" in v)
    ]
    if not customer_candidates:
        customer_candidates = [
            i for i in range(len(header_lower))
            if i not in (header_map["product_id"], header_map["router_port"], header_map["z_pp_number"])
        ]

    port_candidates = [
        i for i, v in enumerate(header_lower)
        if v and ("port" in v) and (i != router_port_idx)
    ]
    if not port_candidates:
        port_candidates = [
            i for i in range(len(header_lower))
            if i not in (header_map["product_id"], header_map["router_port"], header_map["z_pp_number"])
        ]

    best_customer = _choose_best_col(ws, header_row_idx, customer_candidates, _score_customer_cell)
    best_port = _choose_best_col(ws, header_row_idx, port_candidates, _score_zport_cell)

    if best_customer is None or best_port is None:
        raise ValueError("Konnte Customer/Port Spalten nicht sicher erkennen. Bitte Header prüfen.")

    header_map["customer_name"] = best_customer
    header_map["z_port_label"] = best_port

    # ---- read rows
    out: List[Dict[str, Any]] = []
    empty_streak = 0

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        product_id = _norm(row_vals[header_map["product_id"]])
        router_port = _norm(row_vals[header_map["router_port"]])
        customer_name = _norm(row_vals[header_map["customer_name"]])
        z_pp_number = _norm(row_vals[header_map["z_pp_number"]])
        z_port_label = _norm_port_label(row_vals[header_map["z_port_label"]])

        eqx_serial = ""
        if header_map.get("eqx_serial", -1) != -1:
            eqx_serial = _norm(row_vals[header_map["eqx_serial"]])

        if not (product_id or router_port or customer_name or z_pp_number or z_port_label):
            empty_streak += 1
            if empty_streak >= 12:
                break
            continue
        empty_streak = 0

        if _lower(product_id).startswith("dbs") or _lower(product_id).startswith("product"):
            continue

        out.append({
            "_excel_row": r,
            "product_id": product_id,
            "router_port": router_port,
            "customer_name": customer_name,
            "z_pp_number": z_pp_number,
            "z_port_label": z_port_label,
            "eqx_serial": eqx_serial,
        })

    return out


# ============================================================
# DB resolvers
# ============================================================

def parse_router_port(router_port_raw: str) -> Tuple[str, str]:
    v = _norm(router_port_raw)
    m = re.match(r"^([A-Za-z0-9]+)\-(.+)$", v)
    if not m:
        raise ValueError(f"Router Port Format ungültig: {v} (erwartet 'RFRAxxxx-E1/12')")
    return m.group(1).strip(), m.group(2).strip()

def router_iface_to_precable_port(iface: str) -> str:
    p = _norm(iface).upper()
    if p.startswith("ETH"):
        return p
    if p.startswith("E"):
        return "ETH" + p[1:]
    return p

def resolve_a_side_from_router(db: Session, router_port_raw: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    try:
        dev, iface = parse_router_port(router_port_raw)
    except ValueError as e:
        return None, str(e)

    sw_port = router_iface_to_precable_port(iface)

    row = db.execute(text("""
        SELECT patchpanel_id, patchpanel_port, room, switch_name, switch_port
        FROM pre_cabled_links
        WHERE switch_name = :sw
          AND upper(switch_port) = :p
        LIMIT 1;
    """), {"sw": dev, "p": sw_port}).mappings().first()

    if not row:
        return None, f"Kein Precable gefunden für {dev} {sw_port}"

    return {
        "router_device": dev,
        "router_interface": iface,
        "switch_name": row.get("switch_name"),
        "switch_port": row.get("switch_port"),
        "a_pp": row.get("patchpanel_id"),
        "a_port": row.get("patchpanel_port"),
        "a_room": row.get("room"),
    }, None

def resolve_z_port_basic(db: Session, pp_number: str, port_label: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    pp_number = _norm(pp_number)
    port_label_norm = _norm_port_label(port_label)

    pp = db.execute(text("""
        SELECT id, pp_number, room, instance_id, rack_label, rack_unit
        FROM patchpanel_instances
        WHERE pp_number = :pp
        LIMIT 1;
    """), {"pp": pp_number}).mappings().first()

    if not pp:
        return None, f"Z-PP nicht gefunden (pp_number={pp_number})"

    # Compare normalized port labels (ignore spaces/case)
    port = db.execute(text("""
        SELECT id, patchpanel_id, port_label, status, connected_to
        FROM patchpanel_ports
        WHERE patchpanel_id = :pid
          AND upper(replace(port_label,' ','')) = :pl
        LIMIT 1;
    """), {"pid": int(pp["id"]), "pl": port_label_norm}).mappings().first()

    if not port:
        return None, f"Z-Port nicht gefunden (pp_number={pp_number}, port={port_label_norm})"

    st = _lower(port.get("status"))
    if st not in ("free", "available"):
        return None, f"Z-Port nicht verfügbar (status={port.get('status')}, connected_to={port.get('connected_to')})"

    return {
        "z_pp_id": int(pp["id"]),
        "z_pp_number": pp_number,
        "z_port": port_label_norm,
        "z_room": pp.get("room"),
        "z_instance_id": pp.get("instance_id"),
        "z_rack_label": pp.get("rack_label"),
        "z_rack_unit": pp.get("rack_unit"),
    }, None

def resolve_backbone_for_room(db: Session, dest_room: str, prefer_port_label: Optional[str] = None) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """
    ✅ NEU: prefer_port_label wird als "Kassette" interpretiert (A/B/C/D),
    nicht mehr als Customer-Port Label.
    Wir suchen im Zielraum (base_room) einen freien Backbone OUT Port mit gleicher Kassette.
    """
    dest = base_room(dest_room)
    prefer_raw = _norm(prefer_port_label)

    prefer_letter = ""
    if prefer_raw:
        # if user already passes 'B'
        if re.fullmatch(r"[A-Za-z]", prefer_raw):
            prefer_letter = prefer_raw.upper()
        else:
            # if user passes '5B6' etc
            prefer_letter = _extract_cassette_letter(prefer_raw)

    # 1) Try prefer cassette letter
    if prefer_letter:
        row = db.execute(text(r"""
            SELECT
              i.instance_id AS bb_out_pp,
              p.port_label AS bb_out_port,
              p.peer_instance_id AS bb_in_pp,
              p.peer_port_label AS bb_in_port
            FROM patchpanel_ports p
            JOIN patchpanel_instances i ON i.id = p.patchpanel_id
            WHERE i.instance_id ~ '^[0-9]+\.[0-9]+'
              AND substring(i.instance_id from '^[0-9]+\.[0-9]+') = :dest
              AND (p.connected_to IS NULL OR p.connected_to = '')
              AND coalesce(lower(p.status), '') <> 'locked'
              AND substring(upper(replace(p.port_label,' ','')) from '[A-Z]') = :letter
              AND p.peer_instance_id IS NOT NULL
              AND p.peer_port_label IS NOT NULL
            ORDER BY
              i.id DESC,
              coalesce(nullif(substring(upper(replace(p.port_label,' ','')) from '^[0-9]+'), '')::int, 9999),
              coalesce(nullif(substring(upper(replace(p.port_label,' ','')) from '[0-9]+$'), '')::int, 9999)
            LIMIT 1;
        """), {"dest": dest, "letter": prefer_letter}).mappings().first()

        if row:
            return dict(row), None

    # 2) Fallback: any free backbone out port
    row2 = db.execute(text(r"""
        SELECT
          i.instance_id AS bb_out_pp,
          p.port_label AS bb_out_port,
          p.peer_instance_id AS bb_in_pp,
          p.peer_port_label AS bb_in_port
        FROM patchpanel_ports p
        JOIN patchpanel_instances i ON i.id = p.patchpanel_id
        WHERE i.instance_id ~ '^[0-9]+\.[0-9]+'
          AND substring(i.instance_id from '^[0-9]+\.[0-9]+') = :dest
          AND (p.connected_to IS NULL OR p.connected_to = '')
          AND coalesce(lower(p.status), '') <> 'locked'
          AND p.peer_instance_id IS NOT NULL
          AND p.peer_port_label IS NOT NULL
        ORDER BY i.id DESC, p.port_label
        LIMIT 1;
    """), {"dest": dest}).mappings().first()

    if not row2:
        return None, f"Kein freier Backbone OUT Port gefunden für Raum {dest}"

    return dict(row2), None



def has_column(db: Session, table: str, column: str, schema: str = "public") -> bool:
    row = db.execute(
        text(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = :schema
              AND table_name = :table
              AND column_name = :column
            LIMIT 1
            """
        ),
        {"schema": schema, "table": table, "column": column},
    ).first()
    return row is not None

def serial_exists(db: Session, serial: str) -> bool:
    s = _norm(serial)
    if not s:
        return False
    try:
        row = db.execute(text("""
            SELECT 1
            FROM cross_connects
            WHERE serial = :s
            LIMIT 1;
        """), {"s": s}).first()
        return row is not None
    except SQLAlchemyError:
        db.rollback()
        return False


# ============================================================
# Preview Builder
# ============================================================

def build_preview(db: Session, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    out: List[Dict[str, Any]] = []
    counts = {"total": 0, "ok": 0, "warn": 0, "error": 0}

    seen_serials = set()

    for r in rows:
        counts["total"] += 1
        idx = r.get("_excel_row")
        item: Dict[str, Any] = {"row_index": idx, **r}

        try:
            serial = _norm(r.get("product_id"))
            router_port_raw = _norm(r.get("router_port"))
            customer = _norm(r.get("customer_name"))
            z_pp = _norm(r.get("z_pp_number"))
            z_port = _norm_port_label(r.get("z_port_label"))

            if not serial or not router_port_raw or not customer or not z_pp or not z_port:
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = "Pflichtfelder fehlen (Product ID / Router Port / Customer / HU/PP Z / Port)"
                out.append(item)
                continue

            if serial in seen_serials:
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = f"Serial doppelt in Excel Datei ({serial})"
                out.append(item)
                continue
            seen_serials.add(serial)

            if serial_exists(db, serial):
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = f"Serial existiert schon im System ({serial})"
                out.append(item)
                continue

            item["serial"] = serial
            item["z_port_label"] = z_port  # store normalized

            a_side, a_err = resolve_a_side_from_router(db, router_port_raw)
            if a_err:
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = a_err
                out.append(item)
                continue
            item.update(a_side)

            z_info, z_err = resolve_z_port_basic(db, z_pp, z_port)
            if z_err:
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = z_err
                out.append(item)
                continue
            item.update(z_info)

            # ✅ NEU: Backbone Prefer = Kassette aus A-Side Precable Port (z.B. 5B6 -> B)
            cass = _extract_cassette_letter(item.get("a_port"))
            item["a_cassette"] = cass  # optional fürs Frontend/Debug

            bb, bb_err = resolve_backbone_for_room(db, str(z_info.get("z_room") or ""), prefer_port_label=cass)
            if bb_err:
                counts["error"] += 1
                item["status"] = "ERROR"
                item["message"] = bb_err
                out.append(item)
                continue
            item.update(bb)

            # Normalize naming for technicians (BB IN/OUT swapped originally)
            # Keep preview consistent with UI expectation:
            #   bb_in_*  <-> bb_out_*
            bi_pp, bi_port = item.get("bb_in_pp"), item.get("bb_in_port")
            bo_pp, bo_port = item.get("bb_out_pp"), item.get("bb_out_port")
            item["bb_in_pp"], item["bb_out_pp"] = bo_pp, bi_pp
            item["bb_in_port"], item["bb_out_port"] = bo_port, bi_port

            counts["ok"] += 1
            item["status"] = "OK"
            item["message"] = "OK"
            out.append(item)

        except SQLAlchemyError:
            db.rollback()
            counts["error"] += 1
            item["status"] = "ERROR"
            item["message"] = "DB Fehler"
            out.append(item)

    return {"counts": counts, "rows": out}


# ============================================================
# FastAPI endpoints
# ============================================================

@router.post("/preview", dependencies=[Depends(require_permissions("upload:write"))])
async def import_preview(
    kw: int = Form(...),
    mode: str = Form("install"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        rows = parse_install_rows(content)
        preview = build_preview(db, rows)
        preview["kw"] = int(kw)
        preview["mode"] = _norm(mode) or "install"
        preview["file_name"] = file.filename
        return preview
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/commit", dependencies=[Depends(require_permissions("upload:write"))])
async def import_commit(
    kw: int = Form(...),
    mode: str = Form("install"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Echter Commit:
    - validiert wie preview
    - wenn errors>0 => 409
    - schreibt OK rows als status=planned in cross_connects (alles-oder-nix)
    """
    try:
        content = await file.read()
        rows = parse_install_rows(content)
        preview = build_preview(db, rows)

        mode_norm = _norm(mode) or "install"

        if preview.get("counts", {}).get("error", 0) > 0:
            raise HTTPException(status_code=409, detail="Commit abgebrochen: Preview enthält Errors")

        inserted: List[Dict[str, Any]] = []

        # --- Ensure Job + Column exist (Postgres self-heal, kein separater Migration-Step nötig)
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS public.import_jobs (
                id bigserial PRIMARY KEY,
                kw integer NOT NULL,
                mode text NOT NULL,
                file_name text,
                created_at timestamptz NOT NULL DEFAULT NOW()
            );
        """))

        # Check: job_id exists
        has_job_id = db.execute(text("""
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema='public'
            AND table_name='cross_connects'
            AND column_name='job_id'
            LIMIT 1;
        """)).scalar()

        if not has_job_id:
            raise HTTPException(
                status_code=500,
                detail=(
                    "DB Migration fehlt: Spalte public.cross_connects.job_id existiert nicht. "
                    "Bitte als postgres ausführen: "
                    "ALTER TABLE public.cross_connects ADD COLUMN job_id BIGINT;"
                )
            )

        # Optional: check FK (nur prüfen, nicht erstellen)
        has_fk = db.execute(text("""
            SELECT 1 FROM pg_constraint
            WHERE conname = 'fk_cross_connects_job'
            LIMIT 1;
        """)).scalar()

        # FK ist nice-to-have -> kein Fehler, nur warnen (optional)
        # if not has_fk: pass

    
        # Create Job row for this upload
        job_id = None
        user_id = current_user.get("id") if isinstance(current_user, dict) else None
        try:
            db.rollback()
            with db.begin():
                job_id = db.execute(text("""
                    INSERT INTO public.import_jobs (kw, mode, file_name)
                    VALUES (:kw, :mode, :file)
                    RETURNING id;
                """), {"kw": int(kw), "mode": mode_norm, "file": file.filename}).scalar()
                job_id = int(job_id)
        
                write_audit_log(
                    db,
                    user_id=user_id,
                    action="job_create",
                    entity_type="import_job",
                    entity_id=job_id,
                    details={"kw": int(kw), "mode": mode_norm, "file": file.filename},
                )
                write_audit_log(
                    db,
                    user_id=user_id,
                    action="job_start",
                    entity_type="import_job",
                    entity_id=job_id,
                    details={"rows_total": len(preview.get('rows') or [])},
                )
        
                for r in (preview.get("rows") or []):
                    if (r.get("status") or "").upper() != "OK":
                        continue
        
                    serial = _norm(r.get("serial") or r.get("product_id"))
                    if not serial:
                        raise HTTPException(status_code=400, detail="Commit: serial fehlt in einer OK-Zeile")
        
                    if serial_exists(db, serial):
                        raise HTTPException(status_code=409, detail=f"Commit abgebrochen: serial existiert schon ({serial})")
        
                    params = {
                        "serial": serial,
                        "switch_name": _norm(r.get("switch_name")),
                        "switch_port": _norm(r.get("switch_port")),
                        "a_patchpanel_id": _norm(r.get("a_pp")),
                        "a_port_label": _norm(r.get("a_port")),
                        # DB schema is swapped vs. UI naming, keep consistent with
                        # routers/cross_connects.py normalization:
                        #   DB backbone_out_* stores BB IN
                        #   DB backbone_in_*  stores BB OUT
                        "backbone_out_instance_id": _norm(r.get("bb_in_pp")),
                        "backbone_out_port_label": _norm(r.get("bb_in_port")),
                        "backbone_in_instance_id": _norm(r.get("bb_out_pp")),
                        "backbone_in_port_label": _norm(r.get("bb_out_port")),
                        "customer_patchpanel_id": int(r.get("z_pp_id")),
                        "customer_port_label": _norm_port_label(r.get("z_port")),
                        "status": "planned",
                        "job_id": job_id,
                    }
        
                    has_pid = has_column(db, "cross_connects", "product_id")
        
                    id_col = "product_id" if has_pid else "serial"
        
                    new_id = db.execute(text(f"""
                        INSERT INTO cross_connects (
                            {id_col},
                            switch_name,
                            switch_port,
                            a_patchpanel_id,
                            a_port_label,
                            backbone_out_instance_id,
                            backbone_out_port_label,
                            backbone_in_instance_id,
                            backbone_in_port_label,
                            customer_patchpanel_id,
                            customer_port_label,
                            job_id,
                            status,
                            created_at
                        ) VALUES (
                            :serial,
                            :switch_name,
                            :switch_port,
                            :a_patchpanel_id,
                            :a_port_label,
                            :backbone_out_instance_id,
                            :backbone_out_port_label,
                            :backbone_in_instance_id,
                            :backbone_in_port_label,
                            :customer_patchpanel_id,
                            :customer_port_label,
                            :job_id,
                            :status,
                            NOW()
                        )
                        RETURNING id;
                    """), params).scalar()
        
                    db.execute(text("""
                        UPDATE cross_connects
                        SET status = :st
                        WHERE id = :id;
                    """), {"st": "planned", "id": int(new_id)})
        
                    inserted.append({"id": int(new_id), "serial": serial, "status": "planned"})
        
                write_audit_log(
                    db,
                    user_id=user_id,
                    action="job_finish",
                    entity_type="import_job",
                    entity_id=job_id,
                    details={"inserted": len(inserted)},
                )
        
        except HTTPException as e:
            db.rollback()
            with db.begin():
                write_audit_log(
                    db,
                    user_id=user_id,
                    action="job_error",
                    entity_type="import_job",
                    entity_id=job_id,
                    details={"error": str(e), "kw": int(kw), "mode": mode_norm, "file": file.filename},
                )
            raise
        except SQLAlchemyError as e:
            db.rollback()
            with db.begin():
                write_audit_log(
                    db,
                    user_id=user_id,
                    action="job_error",
                    entity_type="import_job",
                    entity_id=job_id,
                    details={"error": str(e), "kw": int(kw), "mode": mode_norm, "file": file.filename},
                )
            raise HTTPException(status_code=500, detail=f"Commit DB Error: {e.__class__.__name__}: {str(e)}")
        return {
            "ok": True,
            "kw": int(kw),
            "mode": mode_norm,
            "job_id": job_id,
            "inserted_count": len(inserted),
            "inserted": inserted,
            "note": "Commit erfolgreich: planned Einträge in cross_connects erstellt.",
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
from fastapi import Body

# ============================================================
# PP seeding helpers
# ============================================================

_PORT_RE = re.compile(r"^(\d+)\s*([A-Z])\s*(\d+)$", re.IGNORECASE)

def _cassette_of_port(port_label: str) -> Optional[int]:
    m = _PORT_RE.match(_norm(port_label).upper())
    if not m:
        return None
    return int(m.group(1))

def _expand_cassette_ports(cassette: int) -> List[str]:
    # 24 Ports pro Kassette: 1A1..1A6, 1B1.., 1C.., 1D..
    letters = ["A", "B", "C", "D"]
    positions = [1, 2, 3, 4, 5, 6]
    out = []
    for L in letters:
        for p in positions:
            out.append(f"{cassette}{L}{p}")
    return out

def _pp_exists(db: Session, pp_number: str) -> Optional[int]:
    row = db.execute(text("""
        SELECT id
        FROM patchpanel_instances
        WHERE pp_number = :pp
        LIMIT 1;
    """), {"pp": _norm(pp_number)}).mappings().first()
    return int(row["id"]) if row else None

def _insert_pp_instance(
    db: Session,
    pp_number: str,
    room: str,
    instance_id: Optional[str] = None,
    rack_label: Optional[str] = None,
    rack_unit: Optional[int] = None,
) -> int:
    pp_number = _norm(pp_number)
    room = _norm(room)
    instance_id = _norm(instance_id) or f"PP:{pp_number}"
    rack_label = _norm(rack_label) or None

    new_id = db.execute(text("""
        INSERT INTO patchpanel_instances (pp_number, room, instance_id, rack_label, rack_unit)
        VALUES (:pp_number, :room, :instance_id, :rack_label, :rack_unit)
        RETURNING id;
    """), {
        "pp_number": pp_number,
        "room": room,
        "instance_id": instance_id,
        "rack_label": rack_label,
        "rack_unit": rack_unit,
    }).scalar()

    return int(new_id)

def _ensure_ports(db: Session, patchpanel_id: int, port_labels: List[str]) -> int:
    # Insert ports if missing (keine doppelt)
    created = 0
    for pl in port_labels:
        pl = _norm(pl).upper()
        if not pl:
            continue

        exists = db.execute(text("""
            SELECT 1
            FROM patchpanel_ports
            WHERE patchpanel_id = :pid
              AND port_label = :pl
            LIMIT 1;
        """), {"pid": int(patchpanel_id), "pl": pl}).first()

        if exists:
            continue

        db.execute(text("""
            INSERT INTO patchpanel_ports (patchpanel_id, port_label, status)
            VALUES (:pid, :pl, 'free');
        """), {"pid": int(patchpanel_id), "pl": pl})
        created += 1

    return created


# ============================================================
# New endpoints: find missing PPs + seed them
# ============================================================

@router.post("/missing-pps", dependencies=[Depends(require_permissions("upload:write"))])
async def import_missing_pps(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Upload Installation Excel -> returns missing Z-PPs with needed cassettes.
    """
    try:
        content = await file.read()
        rows = parse_install_rows(content)

        # Collect ports per pp_number from Excel
        ports_by_pp: Dict[str, set] = {}
        for r in rows:
            ppn = _norm(r.get("z_pp_number"))
            pl = _norm(r.get("z_port_label")).upper()
            if not ppn:
                continue
            ports_by_pp.setdefault(ppn, set())
            if pl:
                ports_by_pp[ppn].add(pl)

        missing = []
        for ppn, ports in ports_by_pp.items():
            if _pp_exists(db, ppn) is not None:
                continue

            cassettes = sorted({c for c in (_cassette_of_port(p) for p in ports) if c is not None})
            # fallback: wenn Ports komisch sind, wenigstens Kassette 1 anlegen
            if not cassettes:
                cassettes = [1]

            missing.append({
                "pp_number": ppn,
                "ports_used": sorted(list(ports)),
                "cassettes_needed": cassettes,
            })

        return {"count": len(missing), "missing": missing}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/seed-pps", dependencies=[Depends(require_permissions("upload:write"))])
def seed_pps(
    payload: List[Dict[str, Any]] = Body(...),
    db: Session = Depends(get_db),
):
    """
    Create patchpanel_instances + ports in bulk.
    Expected JSON items:
      {
        "pp_number": "1373585",
        "room": "5.13",
        "cassettes": [1,2],            # optional
        "instance_id": "PP:1373585",   # optional
        "rack_label": "RACK-01",       # optional
        "rack_unit": 21                # optional
      }
    """
    created_instances = 0
    created_ports_total = 0
    out = []

    try:
        for item in payload:
            ppn = _norm(item.get("pp_number"))
            room = _norm(item.get("room"))
            if not ppn or not room:
                raise HTTPException(status_code=400, detail="seed-pps: pp_number und room sind Pflichtfelder")

            cassettes = item.get("cassettes") or [1]
            # sanitize cassettes
            cassettes = [int(x) for x in cassettes if str(x).strip().isdigit()]
            if not cassettes:
                cassettes = [1]

            # get or create instance
            pid = _pp_exists(db, ppn)
            if pid is None:
                pid = _insert_pp_instance(
                    db=db,
                    pp_number=ppn,
                    room=room,
                    instance_id=item.get("instance_id"),
                    rack_label=item.get("rack_label"),
                    rack_unit=item.get("rack_unit"),
                )
                created_instances += 1

            # ensure ports
            all_ports = []
            for c in sorted(set(cassettes)):
                all_ports.extend(_expand_cassette_ports(int(c)))

            created_ports = _ensure_ports(db, pid, all_ports)
            created_ports_total += created_ports

            out.append({
                "pp_number": ppn,
                "patchpanel_id": int(pid),
                "room": room,
                "cassettes": sorted(set(cassettes)),
                "ports_created": int(created_ports),
            })

        db.commit()
        return {
            "ok": True,
            "instances_created": created_instances,
            "ports_created": created_ports_total,
            "items": out,
        }

    except HTTPException:
        db.rollback()
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"DB Error: {e.__class__.__name__}: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
