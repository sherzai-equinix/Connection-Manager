# routers/cross_connects.py
import io
from datetime import date, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Path
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from config import settings
from database import get_db
from security import get_current_user
from audit import write_audit_log

router = APIRouter(
    prefix=f"{settings.api_prefix}/cross-connects",
    tags=["cross-connects"],
)


# ---------------------------------------------------------------------------
# Backbone IN/OUT normalization
# ---------------------------------------------------------------------------
#
# In the current production DB, backbone_in_* and backbone_out_* were filled
# swapped relative to the UI/business wording used by the technicians.
# Rather than forcing an immediate DB migration, we normalize at the API
# boundary so the frontend always sees the expected meaning:
#
#   API field "backbone_in_*"  == DB column backbone_out_*
#   API field "backbone_out_*" == DB column backbone_in_*
#
# This keeps existing data readable and fixes edits/creates consistently.


def _swap_backbone_fields(item: dict) -> dict:
    """Swap BB IN/OUT fields in an API item dict (in-place safe)."""
    if not item:
        return item
    bi_i = item.get("backbone_in_instance_id")
    bi_p = item.get("backbone_in_port_label")
    bo_i = item.get("backbone_out_instance_id")
    bo_p = item.get("backbone_out_port_label")
    item["backbone_in_instance_id"], item["backbone_out_instance_id"] = bo_i, bi_i
    item["backbone_in_port_label"], item["backbone_out_port_label"] = bo_p, bi_p
    return item


def _swap_backbone_payload(payload: dict) -> dict:
    """Swap BB IN/OUT fields coming from the frontend before writing to DB."""
    if not payload:
        return payload
    has_any = any(
        k in payload
        for k in (
            "backbone_in_instance_id",
            "backbone_in_port_label",
            "backbone_out_instance_id",
            "backbone_out_port_label",
        )
    )
    if not has_any:
        return payload
    # swap values
    bi_i = payload.get("backbone_in_instance_id")
    bi_p = payload.get("backbone_in_port_label")
    bo_i = payload.get("backbone_out_instance_id")
    bo_p = payload.get("backbone_out_port_label")
    payload["backbone_in_instance_id"], payload["backbone_out_instance_id"] = bo_i, bi_i
    payload["backbone_in_port_label"], payload["backbone_out_port_label"] = bo_p, bi_p
    return payload


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def has_column(db: Session, table: str, column: str, schema: str = "public") -> bool:
    row = db.execute(
        text(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = :schema
              AND table_name = :table
              AND column_name = :column
            LIMIT 1
            """
        ),
        {"schema": schema, "table": table, "column": column},
    ).first()
    return row is not None


def _ensure_cc_columns(db: Session) -> None:
    """Ensure optional columns exist (safe to call repeatedly)."""
    db.execute(text("""
        ALTER TABLE public.cross_connects
          ADD COLUMN IF NOT EXISTS tech_comment text,
          ADD COLUMN IF NOT EXISTS updated_at timestamptz;
    """))


def _pending_status_overrides(db: Session) -> dict[int, str]:
    try:
        rows = db.execute(
            text(
                """
                SELECT id, type, status, line_id, line1_id, line2_id, created_at
                FROM public.kw_tasks
                WHERE status IN ('pending_install', 'pending_deinstall', 'pending_move', 'pending_path_move')
                ORDER BY created_at DESC, id DESC
                """
            )
        ).mappings().all()
    except Exception:
        return {}

    overrides: dict[int, str] = {}
    for row in rows:
        t = str(row.get("type") or "").lower()
        status = str(row.get("status") or "").lower()
        if not status:
            continue

        if t in {"deinstall", "move"}:
            line_id = row.get("line_id")
            if line_id is not None:
                overrides.setdefault(int(line_id), status)
            continue

        if t == "path_move":
            line1 = row.get("line1_id")
            line2 = row.get("line2_id")
            if line1 is not None:
                overrides.setdefault(int(line1), status)
            if line2 is not None:
                overrides.setdefault(int(line2), status)

    return overrides

LOCKED_STATUSES_FOR_EDIT = {"active", "pending_serial"}

# Simple workflow for status transitions (V1)
_STATUS_FLOW = {
    "planned": {"review", "in_progress", "troubleshoot", "deinstalled"},
    "review": {"in_progress", "troubleshoot", "deinstalled"},
    "in_progress": {"done", "troubleshoot", "deinstalled"},
    "done": {"pending_serial", "troubleshoot", "deinstalled"},
    "troubleshoot": {"in_progress", "done", "deinstalled"},
    "pending_serial": set(),  # locked by default
    "active": set(),          # locked
    "deinstalled": {"planned", "review"},
}


# ------------------------------------------------------------
# CREATE (pending_serial by default, serial is NULL)
# ------------------------------------------------------------
@router.post("/create")
def create_cross_connect(
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    required = [
        "switch_name", "switch_port",
        "a_patchpanel_id", "a_port_label",
        "backbone_out_instance_id", "backbone_out_port_label",
        "backbone_in_instance_id", "backbone_in_port_label",
        "customer_patchpanel_id", "customer_port_label"
    ]
    for k in required:
        if k not in payload or payload[k] in (None, ""):
            raise HTTPException(status_code=400, detail=f"Missing field: {k}")

    serial = payload.get("serial")
    if isinstance(serial, str):
        serial = serial.strip()
        if serial == "":
            serial = None
    if serial is None:
        status = "pending_serial"
    else:
        status = payload.get("status") or "active"

    # Normalize BB IN/OUT to DB schema
    payload_db = _swap_backbone_payload(dict(payload))
    has_tc = has_column(db, "cross_connects", "tech_comment")
    has_ua = has_column(db, "cross_connects", "updated_at")

    try:
        q = text("""
            INSERT INTO public.cross_connects (
              serial,
              switch_name, switch_port,
              a_patchpanel_id, a_port_label,
              backbone_out_instance_id, backbone_out_port_label,
              backbone_in_instance_id, backbone_in_port_label,
              customer_patchpanel_id, customer_port_label,
              manual_patch_id, pp_connection_id,
              status
            )
            VALUES (
              :serial,
              :switch_name, :switch_port,
              :a_patchpanel_id, :a_port_label,
              :backbone_out_instance_id, :backbone_out_port_label,
              :backbone_in_instance_id, :backbone_in_port_label,
              :customer_patchpanel_id, :customer_port_label,
              :manual_patch_id, :pp_connection_id,
              :status
            )
            RETURNING id, serial, status;
        """)

        with db.begin():
            row = db.execute(
                q,
                {
                    "serial": serial,
                    "switch_name": payload_db["switch_name"],
                    "switch_port": payload_db["switch_port"],
                    "a_patchpanel_id": payload_db["a_patchpanel_id"],
                    "a_port_label": payload_db["a_port_label"],
                    "backbone_out_instance_id": payload_db["backbone_out_instance_id"],
                    "backbone_out_port_label": payload_db["backbone_out_port_label"],
                    "backbone_in_instance_id": payload_db["backbone_in_instance_id"],
                    "backbone_in_port_label": payload_db["backbone_in_port_label"],
                    "customer_patchpanel_id": int(payload_db["customer_patchpanel_id"]),
                    "customer_port_label": payload_db["customer_port_label"],
                    "manual_patch_id": payload_db.get("manual_patch_id"),
                    "pp_connection_id": payload_db.get("pp_connection_id"),
                    "status": status,
                },
            ).fetchone()

            # Return normalized meaning
            item = {"id": row[0], "serial": row[1], "status": row[2]}

            write_audit_log(
                db,
                user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                action="cross_connect_create",
                entity_type="cross_connect",
                entity_id=item["id"],
                details={
                    "status": item.get("status"),
                    "serial": item.get("serial"),
                    "switch_name": payload_db.get("switch_name"),
                    "switch_port": payload_db.get("switch_port"),
                },
            )

        return {"success": True, **item}

    except Exception as e:
        msg = str(e).lower()
        if "duplicate key" in msg or "unique" in msg:
            raise HTTPException(status_code=409, detail="Cross-Connect already exists")
        raise HTTPException(status_code=500, detail=f"Cross-Connect create failed: {str(e)}")



# ------------------------------------------------------------
# SEARCH BY SERIAL (unchanged)
# ------------------------------------------------------------
@router.get("/by-serial")
def get_by_serial(serial: str, db: Session = Depends(get_db)):
    q = text("""
        SELECT cc.*,
               z_pp.instance_id AS customer_patchpanel_instance_id,
               z_pp.room        AS customer_room
        FROM public.cross_connects cc
        LEFT JOIN public.patchpanel_instances z_pp
               ON z_pp.id = cc.customer_patchpanel_id
        WHERE cc.serial = :serial
        LIMIT 1;
    """)
    row = db.execute(q, {"serial": serial}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")

    item = _swap_backbone_fields(dict(row))
    return {"success": True, "data": item}


# ------------------------------------------------------------
# GET BY ID (for edit modal)
# ------------------------------------------------------------
@router.get("/item/{cc_id}")
def get_cross_connect(
    cc_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
):
    row = db.execute(
        text("SELECT * FROM public.cross_connects WHERE id = :id"),
        {"id": cc_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    item = _swap_backbone_fields(dict(row))
    return {"success": True, "item": item}


# ------------------------------------------------------------
# PATCH / UPDATE (Backbone IN/OUT + assign/claim)
# ------------------------------------------------------------
@router.patch("/item/{cc_id}")
def update_cross_connect(
    cc_id: int = Path(..., ge=1),
    payload: dict | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    # Normalize BB IN/OUT to DB schema
    payload_db = _swap_backbone_payload(dict(payload))
    has_tc = has_column(db, "cross_connects", "tech_comment")
    has_ua = has_column(db, "cross_connects", "updated_at")

    # Fetch current
    cur = db.execute(
        text("SELECT id, status, assigned_to FROM public.cross_connects WHERE id = :id"),
        {"id": cc_id},
    ).fetchone()
    if not cur:
        raise HTTPException(status_code=404, detail="Not found")

    current_status = (cur[1] or "").lower()
    if current_status in LOCKED_STATUSES_FOR_EDIT:
        raise HTTPException(status_code=409, detail=f"Locked (status={current_status})")

    # Enforce light locking by assignee (prevents two techs from changing the same line)
    current_assignee = (cur[2] or "").strip() if len(cur) > 2 else ""
    actor = None
    if isinstance(payload_db, dict):
        actor = payload_db.get("current_user") or payload_db.get("actor") or payload_db.get("username")
    if actor is not None:
        actor = str(actor).strip()
    if current_assignee and actor and current_assignee.lower() != actor.lower():
        raise HTTPException(status_code=409, detail=f"Locked by {current_assignee}")

    # Allowed fields (V1)
    # Workflow hardening:
    # - done: only status transitions allowed (no field edits)
    # - troubleshoot: only status transitions + tech_comment allowed
    cur_s = current_status or "planned"

    if cur_s == "done":
        allowed_fields = {"status"}
    elif cur_s == "troubleshoot":
        allowed_fields = {"status", "tech_comment", "assigned_to"}
    else:
        allowed_fields = {
            "backbone_in_instance_id",
            "backbone_in_port_label",
            "backbone_out_instance_id",
            "backbone_out_port_label",
            "customer_patchpanel_id",
            "customer_port_label",
            "status",
            "assigned_to",
            "tech_comment",
        }

    if not has_tc:
        allowed_fields.discard("tech_comment")

    updates = {k: payload_db.get(k) for k in allowed_fields if k in payload_db}
    if not updates:
        raise HTTPException(status_code=400, detail="No editable fields provided")

    # Auto-assign (claim light): if a user is provided and record is unassigned, store it.
    # Frontend may send current_user; we claim only if not assigned yet.
    if "assigned_to" not in updates:
        current_user = actor
        if current_user and not current_assignee:
            updates["assigned_to"] = current_user

    if updates.get("assigned_to"):
        # Normalize
        updates["assigned_to"] = str(updates["assigned_to"]).strip() or None

    # Optional status guard: only allow known statuses if provided
    if "status" in updates and updates["status"] is not None:
        s = str(updates["status"]).strip().lower()
        allowed_status = {
            "active",
            "pending_serial",
            "pending_install",
            "pending_deinstall",
            "pending_move",
            "pending_path_move",
            "planned",
            "review",
            "in_progress",
            "done",
            "troubleshoot",
            "deinstalled",
        }
        if s not in allowed_status:
            raise HTTPException(status_code=400, detail="Invalid status")
        # Never allow setting active here (active is only reached via /assign-serial)
        if s == "active":
            raise HTTPException(status_code=400, detail="Use assign-serial to activate")

        # Validate transition against simple flow
        cur_s = current_status or "planned"
        allowed_next = _STATUS_FLOW.get(cur_s, set())
        if s != cur_s and allowed_next and s not in allowed_next:
            raise HTTPException(status_code=400, detail=f"Invalid transition: {cur_s} -> {s}")

        updates["status"] = s

    # Build SQL dynamically
    set_parts = []
    params = {"id": cc_id}
    for k, v in updates.items():
        if k == "assigned_to":
            set_parts.append("assigned_to = :assigned_to")
            # set assigned_at when assigned_to is set and assigned_at is NULL
            set_parts.append("assigned_at = COALESCE(assigned_at, NOW())")
            params["assigned_to"] = v
        else:
            set_parts.append(f"{k} = :{k}")
            params[k] = v

    # Touch updated_at if column exists
    if has_ua:
        set_parts.append("updated_at = NOW()")

    sql = text(
        "UPDATE public.cross_connects SET " + ", ".join(set_parts) + " WHERE id = :id RETURNING *"
    )

    try:
        db.rollback()
        with db.begin():
            row = db.execute(sql, params).mappings().first()
            item = _swap_backbone_fields(dict(row)) if row else None

            new_status = (item.get("status") or "").lower() if item else current_status
            new_assignee = (item.get("assigned_to") or "").strip() if item else current_assignee

            write_audit_log(
                db,
                user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                action="cross_connect_update",
                entity_type="cross_connect",
                entity_id=cc_id,
                details={"fields": sorted(list(updates.keys()))},
            )

            if new_status and new_status != current_status:
                write_audit_log(
                    db,
                    user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                    action="cross_connect_status_change",
                    entity_type="cross_connect",
                    entity_id=cc_id,
                    details={"from": current_status, "to": new_status},
                )

            if new_assignee != (current_assignee or ""):
                write_audit_log(
                    db,
                    user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                    action="cross_connect_assignment",
                    entity_type="cross_connect",
                    entity_id=cc_id,
                    details={"from": current_assignee or None, "to": new_assignee or None},
                )

        return {"success": True, "item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")


# ------------------------------------------------------------
# LIST + FILTER (status / date range / q / pagination)
# ------------------------------------------------------------
@router.get("/list")
def list_cross_connects(
    status: str = Query(
        "active",
        description="active|pending_serial|pending_install|pending_deinstall|pending_move|pending_path_move|planned|review|in_progress|done|troubleshoot|deinstalled|all",
    ),
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    q: str | None = Query(None, description="free text search"),
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    allowed_status = {
        "active",
        "pending_serial",
        "pending_install",
        "pending_deinstall",
        "pending_move",
        "pending_path_move",
        "planned",
        "review",
        "in_progress",
        "done",
        "troubleshoot",
        "deinstalled",
        "all",
    }
    if status not in allowed_status:
        raise HTTPException(status_code=400, detail="Invalid status")

    where = []
    params: dict[str, Any] = {}

    if date_from:
        where.append("created_at >= :date_from")
        params["date_from"] = date_from

    if date_to:
        # inclusive (bis Ende des Tages)
        params["date_to_plus1"] = date_to + timedelta(days=1)
        where.append("created_at < :date_to_plus1")

    if q and q.strip():
        params["q"] = f"%{q.strip()}%"
        where.append("""
            (
              COALESCE(serial,'') ILIKE :q OR
              COALESCE(switch_name,'') ILIKE :q OR
              COALESCE(switch_port,'') ILIKE :q OR
              COALESCE(a_patchpanel_id,'') ILIKE :q OR
              COALESCE(a_port_label,'') ILIKE :q OR
              COALESCE(backbone_out_instance_id,'') ILIKE :q OR
              COALESCE(backbone_out_port_label,'') ILIKE :q OR
              COALESCE(backbone_in_instance_id,'') ILIKE :q OR
              COALESCE(backbone_in_port_label,'') ILIKE :q OR
              COALESCE(customer_port_label,'') ILIKE :q OR
              COALESCE(system_name,'') ILIKE :q OR
              COALESCE(rack_code,'') ILIKE :q
            )
        """)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    list_q = text(f"""
        SELECT *
        FROM public.cross_connects
        {where_sql}
        ORDER BY created_at DESC
        LIMIT 5000;
    """)

    try:
        rows = db.execute(list_q, params).mappings().all()
        items = [_swap_backbone_fields(dict(r)) for r in rows]
        overrides = _pending_status_overrides(db)
        for item in items:
            line_id = int(item.get("id") or 0)
            if line_id and line_id in overrides and str(item.get("status") or "").lower() != "deinstalled":
                item["status"] = overrides[line_id]

        if status != "all":
            items = [x for x in items if str(x.get("status") or "").lower() == status]

        total = len(items)
        items = items[offset:offset + limit]

        return {"success": True, "total": total, "items": items}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cross-Connect list failed: {str(e)}")


# ------------------------------------------------------------
# EXCEL EXPORT
# ------------------------------------------------------------
@router.get("/export")
def export_cross_connects_xlsx(
    status: str = Query("active"),
    q: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Export cross-connects as styled Excel (same filters as /list)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    allowed_status = {
        "active", "pending_serial", "pending_install", "pending_deinstall",
        "pending_move", "pending_path_move", "planned", "review",
        "in_progress", "done", "troubleshoot", "deinstalled", "all",
    }
    if status not in allowed_status:
        raise HTTPException(status_code=400, detail="Invalid status")

    where = []
    params: dict[str, Any] = {}

    if q and q.strip():
        params["q"] = f"%{q.strip()}%"
        where.append("""(
            COALESCE(serial,'') ILIKE :q OR
            COALESCE(switch_name,'') ILIKE :q OR
            COALESCE(switch_port,'') ILIKE :q OR
            COALESCE(a_patchpanel_id,'') ILIKE :q OR
            COALESCE(backbone_out_instance_id,'') ILIKE :q OR
            COALESCE(backbone_in_instance_id,'') ILIKE :q OR
            COALESCE(customer_port_label,'') ILIKE :q OR
            COALESCE(system_name,'') ILIKE :q OR
            COALESCE(rack_code,'') ILIKE :q
        )""")

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    rows = db.execute(
        text(f"SELECT * FROM public.cross_connects {where_sql} ORDER BY created_at DESC"),
        params,
    ).mappings().all()

    items = [_swap_backbone_fields(dict(r)) for r in rows]
    overrides = _pending_status_overrides(db)
    for item in items:
        line_id = int(item.get("id") or 0)
        if line_id and line_id in overrides and str(item.get("status") or "").lower() != "deinstalled":
            item["status"] = overrides[line_id]

    if status != "all":
        items = [x for x in items if str(x.get("status") or "").lower() == status]

    # ── Build workbook ──
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    active_fill = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")
    deinstalled_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
    pending_fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")

    def _safe(v):
        if isinstance(v, datetime):
            return v.replace(tzinfo=None) if v.tzinfo else v
        return v

    headers = [
        "ID", "Serial", "Product ID", "Status",
        "Switch Name", "Switch Port",
        "A Patchpanel", "A Port",
        "BB IN PP", "BB IN Port",
        "BB OUT PP", "BB OUT Port",
        "Z Patchpanel", "Z Port",
        "Kunde", "Rack",
        "Tech Kommentar",
        "Erstellt", "Aktualisiert",
    ]

    ws = wb.active
    ws.title = "Cross Connects"

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for item in items:
        ws.append([
            _safe(item.get("id")),
            item.get("serial") or "-",
            item.get("product_id") or "-",
            str(item.get("status") or "-"),
            item.get("switch_name") or "-",
            item.get("switch_port") or "-",
            item.get("a_patchpanel_id") or "-",
            item.get("a_port_label") or "-",
            item.get("backbone_in_instance_id") or "-",
            item.get("backbone_in_port_label") or "-",
            item.get("backbone_out_instance_id") or "-",
            item.get("backbone_out_port_label") or "-",
            item.get("customer_patchpanel_id") or "-",
            item.get("customer_port_label") or "-",
            item.get("system_name") or "-",
            item.get("rack_code") or "-",
            item.get("tech_comment") or "-",
            _safe(item.get("created_at")),
            _safe(item.get("updated_at")),
        ])

    # Style data rows
    status_col = headers.index("Status") + 1
    for ri in range(2, ws.max_row + 1):
        s_val = str(ws.cell(row=ri, column=status_col).value or "").lower()
        fill = None
        if s_val == "active":
            fill = active_fill
        elif s_val == "deinstalled":
            fill = deinstalled_fill
        elif "pending" in s_val:
            fill = pending_fill
        for cell in ws[ri]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Zusammenfassung", 0)
    ws_sum.append(["Cross Connects Backup"])
    ws_sum["A1"].font = Font(bold=True, size=16)
    ws_sum.append([])
    ws_sum.append(["Exportiert am", _safe(datetime.now().replace(microsecond=0))])
    ws_sum.append(["Filter", status.capitalize()])
    ws_sum.append([])
    ws_sum.append(["Status", "Anzahl"])
    for cell in ws_sum[6]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    status_counts: dict[str, int] = {}
    for item in items:
        s = str(item.get("status") or "unknown").lower()
        status_counts[s] = status_counts.get(s, 0) + 1
    for s, cnt in sorted(status_counts.items()):
        ws_sum.append([s.capitalize(), cnt])
    ws_sum.append([])
    ws_sum.append(["Gesamt", len(items)])
    ws_sum[f"A{ws_sum.max_row}"].font = Font(bold=True)
    ws_sum[f"B{ws_sum.max_row}"].font = Font(bold=True)

    for ri in range(3, ws_sum.max_row + 1):
        for cell in ws_sum[ri]:
            cell.border = thin_border
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 20

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"Cross_Connects_Backup_{today}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ------------------------------------------------------------
# ASSIGN SERIAL (pending_serial -> active, then locked)
# ------------------------------------------------------------
@router.patch("/{cc_id}/assign-serial")
def assign_serial(
    cc_id: int = Path(...),
    payload: dict = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not payload or "serial" not in payload:
        raise HTTPException(status_code=400, detail="Missing field: serial")

    serial = payload.get("serial")
    serial = serial.strip() if isinstance(serial, str) else ""
    if not serial:
        raise HTTPException(status_code=400, detail="serial required")

    has_sn = has_column(db, "cross_connects", "serial_number")
    try:
        db.rollback()
        with db.begin():
            # 1) check exists + status
            row = db.execute(
                text("SELECT id, status FROM public.cross_connects WHERE id = :id"),
                {"id": cc_id},
            ).fetchone()

            if not row:
                raise HTTPException(status_code=404, detail="Not found")

            current_status = row[1]
            if current_status != "pending_serial":
                raise HTTPException(status_code=409, detail="Locked (not pending_serial)")

            # 2) update (unique serial handled by DB index)
            field = "serial_number" if has_sn else "serial"
            upd = db.execute(
                text(
                    f"""
                    UPDATE public.cross_connects
                    SET {field} = :serial,
                        status = 'active'
                    WHERE id = :id
                      AND status = 'pending_serial'
                    RETURNING id, {field}, status;
                    """
                ),
                {"id": cc_id, "serial": serial},
            ).fetchone()

            if not upd:
                raise HTTPException(status_code=409, detail="Locked")

            write_audit_log(
                db,
                user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                action="cross_connect_serial_assign",
                entity_type="cross_connect",
                entity_id=cc_id,
                details={"serial": upd[1]},
            )
            write_audit_log(
                db,
                user_id=current_user.get("id") if isinstance(current_user, dict) else None,
                action="cross_connect_status_change",
                entity_type="cross_connect",
                entity_id=cc_id,
                details={"from": "pending_serial", "to": "active"},
            )

        return {"success": True, "id": upd[0], "serial": upd[1], "status": upd[2]}

    except HTTPException:
        raise
    except Exception as e:
        msg = str(e).lower()
        if "duplicate key" in msg or "unique" in msg:
            raise HTTPException(status_code=409, detail="Serial already exists")
        raise HTTPException(status_code=500, detail=f"Assign serial failed: {str(e)}")
