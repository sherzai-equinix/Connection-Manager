"""routers/troubleshooting.py

API-Endpunkte fuer Troubleshooting:
  - Aktive Leitung per Serial Number suchen
  - BB IN / BB OUT aktualisieren (nur diese Felder)
  - Troubleshooting-Log abrufen
  - Excel-Report herunterladen
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from config import settings
from database import get_db
from security import get_current_user
from audit import write_audit_log

router = APIRouter(
    prefix=f"{settings.api_prefix}/troubleshooting",
    tags=["troubleshooting"],
)

# ---------------------------------------------------------------------------
# Backbone IN/OUT normalisation (same swap logic as cross_connects.py)
# DB backbone_in_* == API backbone_out_* and vice versa
# ---------------------------------------------------------------------------

def _swap_bb(item: dict) -> dict:
    """Swap BB IN/OUT fields from DB representation to API representation."""
    if not item:
        return item
    bi_i = item.get("backbone_in_instance_id")
    bi_p = item.get("backbone_in_port_label")
    bo_i = item.get("backbone_out_instance_id")
    bo_p = item.get("backbone_out_port_label")
    item["backbone_in_instance_id"] = bo_i
    item["backbone_out_instance_id"] = bi_i
    item["backbone_in_port_label"] = bo_p
    item["backbone_out_port_label"] = bi_p
    return item


def _swap_bb_payload(payload: dict) -> dict:
    """Swap BB IN/OUT from API (frontend) to DB representation before writing."""
    bi_i = payload.get("backbone_in_instance_id")
    bi_p = payload.get("backbone_in_port_label")
    bo_i = payload.get("backbone_out_instance_id")
    bo_p = payload.get("backbone_out_port_label")
    payload["backbone_in_instance_id"] = bo_i
    payload["backbone_out_instance_id"] = bi_i
    payload["backbone_in_port_label"] = bo_p
    payload["backbone_out_port_label"] = bi_p
    return payload


# ---------------------------------------------------------------------------
# Ensure troubleshooting_log table exists
# ---------------------------------------------------------------------------
_TABLE_ENSURED = False


def _ensure_table(db: Session) -> None:
    global _TABLE_ENSURED
    if _TABLE_ENSURED:
        return
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS public.troubleshooting_log (
            id               BIGSERIAL PRIMARY KEY,
            cross_connect_id BIGINT NOT NULL,
            serial_number    TEXT NOT NULL,
            troubleshoot_type TEXT NOT NULL,
            ticket_number    TEXT,
            note             TEXT,
            performed_by     TEXT NOT NULL,
            performed_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            old_bb_in_pp     TEXT,
            old_bb_in_port   TEXT,
            old_bb_out_pp    TEXT,
            old_bb_out_port  TEXT,
            new_bb_in_pp     TEXT,
            new_bb_in_port   TEXT,
            new_bb_out_pp    TEXT,
            new_bb_out_port  TEXT
        );
    """))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_ts_log_serial ON public.troubleshooting_log(serial_number)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_ts_log_performed_at ON public.troubleshooting_log(performed_at DESC)"))
    # Work-lines: lines a technician has queued for troubleshooting (persists across sessions)
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS public.troubleshooting_worklines (
            id               BIGSERIAL PRIMARY KEY,
            cross_connect_id BIGINT NOT NULL,
            serial_number    TEXT NOT NULL,
            troubleshoot_type TEXT NOT NULL,
            ticket_number    TEXT,
            note             TEXT,
            created_by       TEXT NOT NULL,
            created_at       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            cc_data          JSONB
        );
    """))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_ts_wl_user ON public.troubleshooting_worklines(created_by)"))
    db.commit()
    _TABLE_ENSURED = True


# ---------------------------------------------------------------------------
# 1) Search active cross-connect by serial_number
# ---------------------------------------------------------------------------
@router.get("/search")
def search_active_line(
    serial: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Find an active cross-connect by serial (or serial_number)."""
    row = db.execute(
        text("""
            SELECT cc.*,
                   z_pp.instance_id AS customer_patchpanel_instance_id,
                   z_pp.room        AS customer_room
            FROM public.cross_connects cc
            LEFT JOIN public.patchpanel_instances z_pp
                   ON z_pp.id = cc.customer_patchpanel_id
            WHERE cc.status = 'active'
              AND (cc.serial = :serial OR cc.serial_number = :serial)
            LIMIT 1;
        """),
        {"serial": serial.strip()},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Keine aktive Leitung mit dieser Serial Number gefunden.")

    item = _swap_bb(dict(row))
    return {"success": True, "data": item}


# ---------------------------------------------------------------------------
# 2) Update BB IN / BB OUT on an active line
# ---------------------------------------------------------------------------
@router.patch("/update-bb/{cc_id}")
def update_backbone(
    cc_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Update only BB IN / BB OUT fields on an active cross-connect.
    Required payload keys:
      - backbone_in_instance_id, backbone_in_port_label
      - backbone_out_instance_id, backbone_out_port_label
      - troubleshoot_type: 'ticket' | 'normal'
      - ticket_number (if type=ticket)
      - note (if type=normal)
    """
    _ensure_table(db)

    required = [
        "backbone_in_instance_id", "backbone_in_port_label",
        "backbone_out_instance_id", "backbone_out_port_label",
        "troubleshoot_type",
    ]
    for k in required:
        if k not in payload or payload[k] in (None, ""):
            raise HTTPException(status_code=400, detail=f"Feld fehlt: {k}")

    ts_type = payload["troubleshoot_type"]
    if ts_type not in ("ticket", "normal"):
        raise HTTPException(status_code=400, detail="troubleshoot_type muss 'ticket' oder 'normal' sein.")

    if ts_type == "ticket" and not payload.get("ticket_number", "").strip():
        raise HTTPException(status_code=400, detail="Troubleticket-Nummer ist erforderlich.")

    # Fetch current line
    cur = db.execute(
        text("SELECT * FROM public.cross_connects WHERE id = :id"),
        {"id": cc_id},
    ).mappings().first()

    if not cur:
        raise HTTPException(status_code=404, detail="Leitung nicht gefunden.")
    if (cur.get("status") or "").lower() != "active":
        raise HTTPException(status_code=409, detail="Nur aktive Leitungen koennen per Troubleshooting bearbeitet werden.")

    cur_dict = _swap_bb(dict(cur))

    # Old BB values (API perspective)
    old_bb_in_pp = cur_dict.get("backbone_in_instance_id") or ""
    old_bb_in_port = cur_dict.get("backbone_in_port_label") or ""
    old_bb_out_pp = cur_dict.get("backbone_out_instance_id") or ""
    old_bb_out_port = cur_dict.get("backbone_out_port_label") or ""

    # New BB values (from payload, API perspective)
    new_bb_in_pp = str(payload["backbone_in_instance_id"]).strip()
    new_bb_in_port = str(payload["backbone_in_port_label"]).strip()
    new_bb_out_pp = str(payload["backbone_out_instance_id"]).strip()
    new_bb_out_port = str(payload["backbone_out_port_label"]).strip()

    # Swap to DB perspective for writing
    db_payload = _swap_bb_payload({
        "backbone_in_instance_id": new_bb_in_pp,
        "backbone_in_port_label": new_bb_in_port,
        "backbone_out_instance_id": new_bb_out_pp,
        "backbone_out_port_label": new_bb_out_port,
    })

    username = "unknown"
    if isinstance(current_user, dict):
        username = current_user.get("username") or current_user.get("sub") or "unknown"

    now = datetime.now(timezone.utc)

    try:
        # Update the cross-connect BB fields
        db.execute(
            text("""
                UPDATE public.cross_connects
                SET backbone_in_instance_id  = :bi_i,
                    backbone_in_port_label   = :bi_p,
                    backbone_out_instance_id = :bo_i,
                    backbone_out_port_label  = :bo_p,
                    updated_at               = :now
                WHERE id = :id
            """),
            {
                "bi_i": db_payload["backbone_in_instance_id"],
                "bi_p": db_payload["backbone_in_port_label"],
                "bo_i": db_payload["backbone_out_instance_id"],
                "bo_p": db_payload["backbone_out_port_label"],
                "now": now,
                "id": cc_id,
            },
        )

        # Write troubleshooting log entry
        db.execute(
            text("""
                INSERT INTO public.troubleshooting_log (
                    cross_connect_id, serial_number, troubleshoot_type,
                    ticket_number, note, performed_by, performed_at,
                    old_bb_in_pp, old_bb_in_port, old_bb_out_pp, old_bb_out_port,
                    new_bb_in_pp, new_bb_in_port, new_bb_out_pp, new_bb_out_port
                ) VALUES (
                    :cc_id, :serial, :ts_type,
                    :ticket, :note, :user, :now,
                    :old_bi_pp, :old_bi_port, :old_bo_pp, :old_bo_port,
                    :new_bi_pp, :new_bi_port, :new_bo_pp, :new_bo_port
                )
            """),
            {
                "cc_id": cc_id,
                "serial": cur_dict.get("serial") or cur_dict.get("serial_number") or "",
                "ts_type": ts_type,
                "ticket": payload.get("ticket_number", "").strip() if ts_type == "ticket" else None,
                "note": payload.get("note", "").strip() if ts_type == "normal" else None,
                "user": username,
                "now": now,
                "old_bi_pp": old_bb_in_pp,
                "old_bi_port": old_bb_in_port,
                "old_bo_pp": old_bb_out_pp,
                "old_bo_port": old_bb_out_port,
                "new_bi_pp": new_bb_in_pp,
                "new_bi_port": new_bb_in_port,
                "new_bo_pp": new_bb_out_pp,
                "new_bo_port": new_bb_out_port,
            },
        )

        db.commit()

        # Audit log
        write_audit_log(
            db,
            user_id=current_user.get("id") if isinstance(current_user, dict) else None,
            action="troubleshooting_bb_update",
            entity_type="cross_connect",
            entity_id=cc_id,
            details={
                "troubleshoot_type": ts_type,
                "ticket_number": payload.get("ticket_number"),
                "note": payload.get("note"),
                "old": {
                    "bb_in_pp": old_bb_in_pp,
                    "bb_in_port": old_bb_in_port,
                    "bb_out_pp": old_bb_out_pp,
                    "bb_out_port": old_bb_out_port,
                },
                "new": {
                    "bb_in_pp": new_bb_in_pp,
                    "bb_in_port": new_bb_in_port,
                    "bb_out_pp": new_bb_out_pp,
                    "bb_out_port": new_bb_out_port,
                },
            },
        )

        return {"success": True, "message": "BB-Weg erfolgreich aktualisiert."}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Fehler beim Speichern: {str(e)}")


# ---------------------------------------------------------------------------
# 3) Get troubleshooting log entries
# ---------------------------------------------------------------------------
@router.get("/log")
def get_troubleshooting_log(
    limit: int = Query(200, ge=1, le=5000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_table(db)
    rows = db.execute(
        text("""
            SELECT * FROM public.troubleshooting_log
            ORDER BY performed_at DESC
            LIMIT :limit
        """),
        {"limit": limit},
    ).mappings().all()
    return {"success": True, "data": [dict(r) for r in rows]}


# ---------------------------------------------------------------------------
# 4) Excel report download
# ---------------------------------------------------------------------------
@router.get("/report")
def download_troubleshooting_report(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_table(db)

    rows = db.execute(
        text("""
            SELECT * FROM public.troubleshooting_log
            ORDER BY performed_at DESC
        """)
    ).mappings().all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nicht installiert.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Troubleshooting Report"

    headers = [
        "Datum", "Uhrzeit", "Durchgefuehrt von", "Typ",
        "Ticket-Nr / Notiz", "Serial Number",
        "Alter BB IN PP", "Alter BB IN Port",
        "Alter BB OUT PP", "Alter BB OUT Port",
        "Neuer BB IN PP", "Neuer BB IN Port",
        "Neuer BB OUT PP", "Neuer BB OUT Port",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        performed_at = row.get("performed_at")
        datum = ""
        uhrzeit = ""
        if performed_at:
            if isinstance(performed_at, str):
                try:
                    performed_at = datetime.fromisoformat(performed_at)
                except Exception:
                    pass
            if isinstance(performed_at, datetime):
                datum = performed_at.strftime("%d.%m.%Y")
                uhrzeit = performed_at.strftime("%H:%M:%S")

        ts_type = row.get("troubleshoot_type", "")
        typ_display = "Troubleticket" if ts_type == "ticket" else "Normales Troubleshooting"
        ticket_or_note = row.get("ticket_number") or row.get("note") or ""

        values = [
            datum, uhrzeit,
            row.get("performed_by", ""),
            typ_display,
            ticket_or_note,
            row.get("serial_number", ""),
            row.get("old_bb_in_pp", ""),
            row.get("old_bb_in_port", ""),
            row.get("old_bb_out_pp", ""),
            row.get("old_bb_out_port", ""),
            row.get("new_bb_in_pp", ""),
            row.get("new_bb_in_port", ""),
            row.get("new_bb_out_pp", ""),
            row.get("new_bb_out_port", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Troubleshooting_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# 5) Worklines CRUD — persist queued lines in DB
# ---------------------------------------------------------------------------

@router.get("/worklines")
def get_worklines(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_table(db)
    username = current_user.get("username") or current_user.get("sub") or "unknown"
    rows = db.execute(
        text("SELECT * FROM public.troubleshooting_worklines WHERE created_by = :u ORDER BY created_at"),
        {"u": username},
    ).mappings().all()
    return {"success": True, "items": [dict(r) for r in rows]}


@router.get("/worklines/all")
def get_all_worklines(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Return all worklines from all users (for dashboard widget)."""
    _ensure_table(db)
    rows = db.execute(
        text("SELECT id, cross_connect_id, serial_number, troubleshoot_type, ticket_number, note, created_by, created_at FROM public.troubleshooting_worklines ORDER BY created_at DESC"),
    ).mappings().all()
    return {"success": True, "items": [dict(r) for r in rows]}


@router.post("/worklines")
def add_workline(
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    import json as _json
    import traceback as _tb

    _ensure_table(db)
    username = current_user.get("username") or current_user.get("sub") or "unknown"
    cc_id = payload.get("cross_connect_id")
    serial = payload.get("serial_number", "")
    ts_type = payload.get("troubleshoot_type", "normal")
    ticket_nr = payload.get("ticket_number", "")
    note = payload.get("note", "")
    cc_data = payload.get("cc_data", {})

    if not cc_id:
        raise HTTPException(status_code=400, detail="cross_connect_id fehlt.")

    try:
        # Prevent duplicates per user
        existing = db.execute(
            text("SELECT id FROM public.troubleshooting_worklines WHERE created_by = :u AND cross_connect_id = :cc"),
            {"u": username, "cc": cc_id},
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="Diese Leitung ist bereits in der Arbeitsliste.")

        cc_data_str = _json.dumps(cc_data) if cc_data else "{}"
        db.execute(
            text("""
                INSERT INTO public.troubleshooting_worklines
                    (cross_connect_id, serial_number, troubleshoot_type, ticket_number, note, created_by, cc_data)
                VALUES (:cc, :serial, :ts_type, :ticket, :note, :user, cast(:cc_data as jsonb))
            """),
            {
                "cc": int(cc_id),
                "serial": serial or "",
                "ts_type": ts_type or "normal",
                "ticket": ticket_nr or None,
                "note": note or None,
                "user": username,
                "cc_data": cc_data_str,
            },
        )
        db.commit()
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"[TS] add_workline ERROR: {e}\n{_tb.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Workline speichern fehlgeschlagen: {str(e)}")


@router.delete("/worklines/{cc_id}")
def remove_workline(
    cc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_table(db)
    username = current_user.get("username") or current_user.get("sub") or "unknown"
    db.execute(
        text("DELETE FROM public.troubleshooting_worklines WHERE created_by = :u AND cross_connect_id = :cc"),
        {"u": username, "cc": cc_id},
    )
    db.commit()
    return {"success": True}
