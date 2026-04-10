from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Path, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import text
from sqlalchemy.orm import Session

from config import settings
from database import get_db
from security import get_current_user


router = APIRouter(prefix=settings.api_prefix, tags=["kw-flow"])


PLAN_STATUS_ALLOWED = {"open", "locked", "completed"}
PLAN_EDIT_BLOCKED = {"locked", "completed"}
CHANGE_TYPE_ALLOWED = {"NEW_INSTALL", "LINE_MOVE", "PATH_MOVE", "DEINSTALL"}
CHANGE_STATUS_ALLOWED = {"planned", "in_progress", "done", "canceled"}
CHANGE_APPLY_ALLOWED = {"planned", "in_progress"}


def _swap_backbone_fields(item: dict[str, Any] | None) -> dict[str, Any] | None:
    if not item:
        return item
    bi_i = item.get("backbone_in_instance_id")
    bi_p = item.get("backbone_in_port_label")
    bo_i = item.get("backbone_out_instance_id")
    bo_p = item.get("backbone_out_port_label")
    item["backbone_in_instance_id"], item["backbone_out_instance_id"] = bo_i, bi_i
    item["backbone_in_port_label"], item["backbone_out_port_label"] = bo_p, bi_p
    return item


def _swap_backbone_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not payload:
        return payload
    bi_i = payload.get("backbone_in_instance_id")
    bi_p = payload.get("backbone_in_port_label")
    bo_i = payload.get("backbone_out_instance_id")
    bo_p = payload.get("backbone_out_port_label")
    payload["backbone_in_instance_id"], payload["backbone_out_instance_id"] = bo_i, bi_i
    payload["backbone_in_port_label"], payload["backbone_out_port_label"] = bo_p, bi_p
    return payload


def _serialize_json(value: dict[str, Any] | None) -> str:
    return json.dumps(value or {}, ensure_ascii=False, default=str)


def _kw_label(year: int, kw: int) -> str:
    return f"{int(year)}-KW{int(kw):02d}"


def _parse_kw_label(raw: str | None) -> tuple[int, int]:
    value = str(raw or "").strip().upper()
    match = re.match(r"^(\d{4})-?KW(\d{1,2})$", value)
    if not match:
        raise HTTPException(status_code=400, detail="kw must be formatted as YYYY-KWNN")
    year = int(match.group(1))
    kw = int(match.group(2))
    if year < 2000 or year > 2100 or kw < 1 or kw > 53:
        raise HTTPException(status_code=400, detail="Invalid kw value")
    return year, kw


def _normalize_plan_status(raw: str | None) -> str:
    status = str(raw or "").strip().lower()
    if status in PLAN_STATUS_ALLOWED:
        return status
    if status in {"active", "draft"}:
        return "open"
    if status in {"archived", "done"}:
        return "completed"
    return "open"


def _has_column(db: Session, table: str, column: str, schema: str = "public") -> bool:
    try:
        row = db.execute(
            text(
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = :schema
                  AND table_name = :table
                  AND column_name = :column
                LIMIT 1
                """
            ),
            {"schema": schema, "table": table, "column": column},
        ).first()
        return row is not None
    except Exception:
        return False


def _ensure_kw_flow_tables(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS public.kw_plans (
                id BIGSERIAL PRIMARY KEY,
                year INTEGER NOT NULL,
                kw INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'open',
                created_by BIGINT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                completed_at TIMESTAMPTZ NULL,
                archived_at TIMESTAMPTZ NULL,
                CONSTRAINT uq_kw_plans_year_kw UNIQUE (year, kw)
            );
            """
        )
    )
    db.execute(text("ALTER TABLE public.kw_plans ADD COLUMN IF NOT EXISTS status TEXT"))
    db.execute(text("ALTER TABLE public.kw_plans ADD COLUMN IF NOT EXISTS created_by BIGINT NULL"))
    db.execute(text("ALTER TABLE public.kw_plans ADD COLUMN IF NOT EXISTS completed_at TIMESTAMPTZ NULL"))
    db.execute(text("ALTER TABLE public.kw_plans ADD COLUMN IF NOT EXISTS archived_at TIMESTAMPTZ NULL"))
    db.execute(
        text(
            """
            UPDATE public.kw_plans
            SET status = CASE
                WHEN LOWER(COALESCE(status, '')) IN ('open', 'locked', 'completed') THEN LOWER(status)
                WHEN LOWER(COALESCE(status, '')) IN ('active', 'draft') THEN 'open'
                WHEN LOWER(COALESCE(status, '')) IN ('archived', 'done') THEN 'completed'
                ELSE 'open'
            END
            """
        )
    )
    try:
        db.execute(text("ALTER TABLE public.kw_plans ALTER COLUMN status SET DEFAULT 'open'"))
        db.execute(text("ALTER TABLE public.kw_plans ALTER COLUMN status SET NOT NULL"))
    except Exception:
        pass
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_kw_plans_status_year_kw ON public.kw_plans(status, year, kw)"))

    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS public.kw_changes (
                id BIGSERIAL PRIMARY KEY,
                kw_plan_id BIGINT NOT NULL REFERENCES public.kw_plans(id) ON DELETE CASCADE,
                type TEXT NOT NULL,
                target_cross_connect_id BIGINT NULL,
                payload_json JSONB NULL,
                status TEXT NOT NULL DEFAULT 'planned',
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                completed_at TIMESTAMPTZ NULL,
                created_by BIGINT NULL,
                completed_by BIGINT NULL
            );
            """
        )
    )
    db.execute(text("ALTER TABLE public.kw_changes ADD COLUMN IF NOT EXISTS completed_at TIMESTAMPTZ NULL"))
    db.execute(text("ALTER TABLE public.kw_changes ADD COLUMN IF NOT EXISTS created_by BIGINT NULL"))
    db.execute(text("ALTER TABLE public.kw_changes ADD COLUMN IF NOT EXISTS completed_by BIGINT NULL"))
    db.execute(
        text(
            """
            UPDATE public.kw_changes
            SET status = CASE
                WHEN LOWER(COALESCE(status, '')) IN ('planned', 'in_progress', 'done', 'canceled') THEN LOWER(status)
                WHEN LOWER(COALESCE(status, '')) = 'cancelled' THEN 'canceled'
                ELSE 'planned'
            END
            """
        )
    )
    try:
        db.execute(text("ALTER TABLE public.kw_changes ALTER COLUMN status SET DEFAULT 'planned'"))
        db.execute(text("ALTER TABLE public.kw_changes ALTER COLUMN status SET NOT NULL"))
    except Exception:
        pass
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_kw_changes_plan ON public.kw_changes(kw_plan_id)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_kw_changes_type_status ON public.kw_changes(type, status)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_kw_changes_target ON public.kw_changes(target_cross_connect_id)"))

    # Archive table for deinstalled cross-connects
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS public.cross_connects_archive (
                id                          BIGSERIAL PRIMARY KEY,
                original_id                 BIGINT NOT NULL,
                serial                      TEXT,
                serial_number               TEXT,
                product_id                  TEXT,
                switch_name                 TEXT,
                switch_port                 TEXT,
                a_patchpanel_id             TEXT,
                a_port_label                TEXT,
                backbone_out_instance_id    TEXT,
                backbone_out_port_label     TEXT,
                backbone_in_instance_id     TEXT,
                backbone_in_port_label      TEXT,
                customer_patchpanel_id      BIGINT,
                customer_port_label         TEXT,
                z_pp_number                 TEXT,
                rack_code                   TEXT,
                system_name                 TEXT,
                customer_id                 BIGINT,
                customer_rack_id            BIGINT,
                customer_location_id        INTEGER,
                job_id                      BIGINT,
                source_audit_line_id        BIGINT,
                status                      TEXT NOT NULL DEFAULT 'deinstalled',
                original_created_at         TIMESTAMPTZ,
                deinstalled_at              TIMESTAMPTZ DEFAULT NOW(),
                deinstalled_by              TEXT,
                reason                      TEXT
            );
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_cc_archive_serial ON public.cross_connects_archive(serial)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_cc_archive_orig_id ON public.cross_connects_archive(original_id)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_cc_archive_system ON public.cross_connects_archive(system_name)"))

    # One-time migration: move any existing deinstalled records into archive
    db.execute(
        text(
            """
            INSERT INTO public.cross_connects_archive (
                original_id, serial, serial_number, product_id,
                switch_name, switch_port,
                a_patchpanel_id, a_port_label,
                backbone_out_instance_id, backbone_out_port_label,
                backbone_in_instance_id, backbone_in_port_label,
                customer_patchpanel_id, customer_port_label,
                z_pp_number, rack_code, system_name,
                customer_id, customer_rack_id, customer_location_id,
                job_id, source_audit_line_id,
                status, original_created_at,
                deinstalled_at, deinstalled_by, reason
            )
            SELECT
                id, serial, serial_number, product_id,
                switch_name, switch_port,
                a_patchpanel_id, a_port_label,
                backbone_out_instance_id, backbone_out_port_label,
                backbone_in_instance_id, backbone_in_port_label,
                customer_patchpanel_id, customer_port_label,
                z_pp_number, rack_code, system_name,
                customer_id, customer_rack_id, customer_location_id,
                job_id, source_audit_line_id,
                'deinstalled', created_at,
                COALESCE(updated_at, NOW()), '', 'Legacy migration'
            FROM public.cross_connects
            WHERE LOWER(COALESCE(status, '')) = 'deinstalled'
              AND id NOT IN (SELECT original_id FROM public.cross_connects_archive)
            """
        )
    )
    # Remove the migrated deinstalled records from active table
    db.execute(
        text(
            """
            DELETE FROM public.cross_connects
            WHERE LOWER(COALESCE(status, '')) = 'deinstalled'
              AND id IN (SELECT original_id FROM public.cross_connects_archive)
            """
        )
    )


def _plan_row_to_out(row: dict[str, Any]) -> dict[str, Any]:
    year = int(row["year"])
    kw = int(row["kw"])
    status = _normalize_plan_status(row.get("status"))
    return {
        "id": int(row["id"]),
        "kw": _kw_label(year, kw),
        "year": year,
        "kw_number": kw,
        "status": status,
        "created_at": row.get("created_at"),
        "created_by": row.get("created_by"),
        "changes_total": int(row.get("changes_total") or 0),
        "open_changes": int(row.get("open_changes") or 0),
    }


# Backbone rooms (A-side) – never a real customer room
_BACKBONE_ROOMS_NORM = {
    'M504S6', '504S6', '54S6', 'M54S6',
    'M513S1', '513S1', '5131', 'M5131',
    '504S6', 'M504S6',
}


def _line_select_sql(db: Session) -> str:
    serial_expr = "COALESCE(cc.serial_number, cc.serial)" if _has_column(db, "cross_connects", "serial_number") else "cc.serial"
    return f"""
        SELECT
            cc.*,
            {serial_expr} AS serial_effective,
            a_pp.room AS a_room,
            a_pp.rack_label AS a_rack,
            z_pp.room AS z_room,
            z_pp.rack_label AS z_rack,
            z_pp.instance_id AS customer_patchpanel_instance_id,
            cust.name AS customer_base_name,
            (
                SELECT cl.room
                FROM public.customer_locations cl
                WHERE cl.customer_id = z_pp.customer_id
                  AND cl.room IS NOT NULL AND cl.room <> ''
                  AND upper(regexp_replace(cl.room, '[^A-Z0-9]', '', 'g'))
                      NOT IN ('M504S6','504S6','54S6','M54S6','M513S1','513S1','5131','M5131')
                ORDER BY cl.id
                LIMIT 1
            ) AS z_customer_room,
            (
                SELECT cc2.system_name
                FROM public.cross_connects cc2
                WHERE cc2.customer_patchpanel_id = cc.customer_patchpanel_id
                  AND cc2.system_name LIKE '%%:%%'
                ORDER BY cc2.id
                LIMIT 1
            ) AS resolved_system_name
        FROM public.cross_connects cc
        LEFT JOIN public.patchpanel_instances a_pp
               ON a_pp.instance_id = cc.a_patchpanel_id
               OR CAST(a_pp.id AS TEXT) = cc.a_patchpanel_id
               OR a_pp.pp_number = cc.a_patchpanel_id
        LEFT JOIN public.patchpanel_instances z_pp
               ON z_pp.id = cc.customer_patchpanel_id
        LEFT JOIN public.customers cust
               ON cust.id = z_pp.customer_id
    """


def _line_by_id(db: Session, line_id: int, for_update: bool = False) -> dict[str, Any] | None:
    sql = _line_select_sql(db) + " WHERE cc.id = :id"
    if for_update:
        sql += " FOR UPDATE OF cc"
    row = db.execute(text(sql), {"id": int(line_id)}).mappings().first()
    if not row:
        return None
    item = dict(row)
    item["serial"] = item.get("serial_effective") or item.get("serial")
    item.pop("serial_effective", None)
    return _swap_backbone_fields(item) or item


def _extract_room_from_system_name(system_name: str | None) -> str:
    """Extract the customer room from system_name pattern.
    E.g. 'FR2:OG-M1A2:OC:Susquehanna' → '1A2'
         'FR2:EG-M5.12:S1:SUSQUEHANNA' → '5.12'
    """
    if not system_name or ":" not in system_name:
        return ""
    m = re.search(r'-?M(\d[A-Z0-9.]+)', system_name, re.IGNORECASE)
    if m:
        return m.group(1)
    return ""


def _is_backbone_room(room: str) -> bool:
    """Check if a room string (e.g. '5.4S6', '1A2') is a backbone room."""
    norm = re.sub(r'[^A-Z0-9]', '', room.upper())
    return norm in _BACKBONE_ROOMS_NORM


def _extract_room_from_bb_instance(bb_in_id: str | None) -> str:
    """Extract the customer room from backbone_in_instance_id.
    DB stores: '1A2/RU41' → room '1A2' (customer side)
               '5.4S6/RU39' → '5.4S6' (backbone side, skip)
    Returns the room only if it's NOT a backbone room.
    """
    if not bb_in_id:
        return ""
    room = bb_in_id.split("/")[0].strip()
    if room and not _is_backbone_room(room):
        return room
    return ""


def _line_to_public(line: dict[str, Any]) -> dict[str, Any]:
    # Z-side PP: prefer the actual instance_id (e.g. PP:0607:1370187)
    z_pp = (
        line.get("customer_patchpanel_instance_id")
        or line.get("z_pp_number")
        or line.get("customer_patchpanel_id")
    )
    # Customer display: prefer system_name with ':' (full format),
    # then resolved_system_name from sibling cross-connect on same PP,
    # otherwise use customer_base_name from the customers table.
    sys_name = line.get("system_name") or ""
    resolved_sn = line.get("resolved_system_name") or ""
    base_name = line.get("customer_base_name") or ""
    if ":" in sys_name:
        customer = sys_name
    elif ":" in resolved_sn:
        customer = resolved_sn
    else:
        customer = sys_name or base_name or line.get("customer_name") or line.get("customer") or "-"
    # Z-Room: prefer room extracted from system_name, then from BB routing,
    # then customer_locations subquery, then PP room
    room_from_sn = _extract_room_from_system_name(sys_name)
    # After _swap_backbone_fields, backbone_out_instance_id = DB backbone_in (customer side)
    room_from_bb = _extract_room_from_bb_instance(line.get("backbone_out_instance_id")) if not room_from_sn else ""
    z_customer_room = room_from_sn or room_from_bb or line.get("z_customer_room") or ""
    z_room = z_customer_room or line.get("z_room") or ""
    item = {
        "id": int(line["id"]),
        "serial": line.get("serial"),
        "customer": customer,
        "system_name": line.get("system_name"),
        "customer_base_name": base_name,
        "switch_name": line.get("switch_name"),
        "switch_port": line.get("switch_port"),
        "a_patchpanel_id": line.get("a_patchpanel_id"),
        "a_port_label": line.get("a_port_label"),
        "customer_patchpanel_id": line.get("customer_patchpanel_id"),
        "customer_patchpanel_instance_id": line.get("customer_patchpanel_instance_id"),
        "customer_port_label": line.get("customer_port_label"),
        "z_pp_number": line.get("z_pp_number"),
        "rack_code": line.get("rack_code"),
        "backbone_in_instance_id": line.get("backbone_in_instance_id"),
        "backbone_in_port_label": line.get("backbone_in_port_label"),
        "backbone_out_instance_id": line.get("backbone_out_instance_id"),
        "backbone_out_port_label": line.get("backbone_out_port_label"),
        "a_room": line.get("a_room"),
        "a_rack": line.get("a_rack"),
        "z_room": z_room,
        "z_rack": line.get("z_rack"),
        "z_customer_room": z_customer_room,
        "a_side": {
            "room": line.get("a_room"),
            "rack": line.get("a_rack"),
            "pp": line.get("a_patchpanel_id"),
            "port": line.get("a_port_label"),
        },
        "z_side": {
            "room": z_room,
            "rack": line.get("z_rack") or line.get("rack_code"),
            "pp": z_pp,
            "port": line.get("customer_port_label"),
        },
        "bb_in": {
            "pp": line.get("backbone_in_instance_id"),
            "port": line.get("backbone_in_port_label"),
        },
        "bb_out": {
            "pp": line.get("backbone_out_instance_id"),
            "port": line.get("backbone_out_port_label"),
        },
        "status": line.get("status"),
    }
    return item


def _assert_plan_writable(plan: dict[str, Any]) -> None:
    status = _normalize_plan_status(plan.get("status"))
    if status in PLAN_EDIT_BLOCKED:
        raise HTTPException(status_code=403, detail=f"KW plan is {status}; changes are blocked")


def _find_plan_by_kw(db: Session, kw_label: str, for_update: bool = False) -> dict[str, Any] | None:
    year, kw = _parse_kw_label(kw_label)
    sql = """
        SELECT id, year, kw, status, created_by, created_at
        FROM public.kw_plans
        WHERE year = :year AND kw = :kw
        LIMIT 1
    """
    if for_update:
        sql += " FOR UPDATE"
    row = db.execute(text(sql), {"year": year, "kw": kw}).mappings().first()
    return dict(row) if row else None


def _find_plan_by_id(db: Session, plan_id: int, for_update: bool = False) -> dict[str, Any] | None:
    sql = """
        SELECT id, year, kw, status, created_by, created_at
        FROM public.kw_plans
        WHERE id = :id
        LIMIT 1
    """
    if for_update:
        sql += " FOR UPDATE"
    row = db.execute(text(sql), {"id": int(plan_id)}).mappings().first()
    return dict(row) if row else None


def _get_user_id(current_user: Any) -> int | None:
    if isinstance(current_user, dict):
        raw = current_user.get("id")
        try:
            return int(raw) if raw is not None else None
        except Exception:
            return None
    return None


def _assert_line_active(line: dict[str, Any] | None, label: str) -> dict[str, Any]:
    if not line:
        raise HTTPException(status_code=404, detail=f"{label} not found")
    if str(line.get("status") or "").lower() != "active":
        raise HTTPException(status_code=409, detail=f"{label} must be active")
    return line


def _validate_change_status(status: str | None) -> str:
    normalized = str(status or "planned").strip().lower()
    if normalized == "cancelled":
        normalized = "canceled"
    if normalized not in CHANGE_STATUS_ALLOWED:
        raise HTTPException(status_code=400, detail="Invalid change status")
    return normalized


def _assert_z_port_free(
    db: Session,
    patchpanel_id: int,
    port_label: str,
    exclude_line_id: int | None = None,
) -> None:
    params: dict[str, Any] = {"pp": int(patchpanel_id), "port": str(port_label).strip()}
    where_extra = ""
    if exclude_line_id:
        params["exclude"] = int(exclude_line_id)
        where_extra = "AND id <> :exclude"
    row = db.execute(
        text(
            """
            SELECT id
            FROM public.cross_connects
            WHERE status = 'active'
              AND customer_patchpanel_id = :pp
              AND customer_port_label = :port
              """
            + where_extra
            + """
            LIMIT 1
            """
        ),
        params,
    ).first()
    if row:
        raise HTTPException(status_code=409, detail="Z-side port already occupied by another active line")


def _assert_field_pair_free(
    db: Session,
    col_instance: str,
    col_port: str,
    instance_value: str,
    port_value: str,
    exclude_line_id: int | None = None,
) -> None:
    params: dict[str, Any] = {"instance": str(instance_value).strip(), "port": str(port_value).strip()}
    where_extra = ""
    if exclude_line_id:
        params["exclude"] = int(exclude_line_id)
        where_extra = "AND id <> :exclude"
    row = db.execute(
        text(
            f"""
            SELECT id
            FROM public.cross_connects
            WHERE status = 'active'
              AND {col_instance} = :instance
              AND {col_port} = :port
              {where_extra}
            LIMIT 1
            """
        ),
        params,
    ).first()
    if row:
        raise HTTPException(status_code=409, detail=f"Port already occupied: {col_instance}/{col_port}")


def _change_row_to_out(row: dict[str, Any]) -> dict[str, Any]:
    payload = row.get("payload_json") or {}
    target = row.get("target_cross_connect_id")
    short = ""
    change_type = str(row.get("type") or "")

    if change_type == "NEW_INSTALL":
        line = payload.get("new_line") or {}
        short = f"Install {line.get('serial') or line.get('product_id') or '-'}"
    elif change_type == "LINE_MOVE":
        old_z = payload.get("old_z") or {}
        new_z = payload.get("new_z") or {}
        short = f"Z {old_z.get('customer_patchpanel_id')}/{old_z.get('customer_port_label')} -> {new_z.get('customer_patchpanel_id')}/{new_z.get('customer_port_label')}"
    elif change_type == "PATH_MOVE":
        short = f"Swap BB {payload.get('line_a_id')} <-> {payload.get('line_b_id')}"
    elif change_type == "DEINSTALL":
        short = f"Deinstall line {target or '-'}"

    return {
        "id": int(row["id"]),
        "kw_plan_id": int(row["kw_plan_id"]),
        "type": change_type,
        "target_cross_connect_id": target,
        "payload_json": payload,
        "status": str(row.get("status") or "planned").lower().replace("cancelled", "canceled"),
        "created_at": row.get("created_at"),
        "completed_at": row.get("completed_at"),
        "created_by": row.get("created_by"),
        "completed_by": row.get("completed_by"),
        "kw": _kw_label(int(row["year"]), int(row["kw"])),
        "short_description": short,
    }


class KwPlanCreateIn(BaseModel):
    kw: str = Field(..., description="YYYY-KWNN")
    status: str | None = Field(default="open")

    @model_validator(mode="after")
    def _check_status(self) -> "KwPlanCreateIn":
        status = _normalize_plan_status(self.status)
        if status not in PLAN_STATUS_ALLOWED:
            raise ValueError("Invalid plan status")
        self.status = status
        return self


class KwPlanPatchIn(BaseModel):
    status: str


class KwChangeCreateIn(BaseModel):
    kw_plan_id: int | None = None
    kw: str | None = None
    type: str
    target_cross_connect_id: int | None = None
    payload_json: dict[str, Any] = Field(default_factory=dict)
    status: str | None = "planned"

    @model_validator(mode="after")
    def _validate_locator(self) -> "KwChangeCreateIn":
        if not self.kw_plan_id and not self.kw:
            raise ValueError("kw_plan_id or kw is required")
        return self


class KwChangePatchIn(BaseModel):
    status: str | None = None
    payload_json: dict[str, Any] | None = None


@router.get("/cross_connects")
def list_cross_connects_minimal(
    status: str = Query("all"),
    q: str | None = Query(None),
    limit: int = Query(200, ge=1, le=2000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)
    allowed = {"all", "active", "deinstalled"}
    normalized_status = str(status or "all").strip().lower()
    if normalized_status not in allowed:
        raise HTTPException(status_code=400, detail="Invalid status filter")

    params: dict[str, Any] = {"limit": int(limit)}
    where = []

    if normalized_status != "all":
        where.append("cc.status = :status")
        params["status"] = normalized_status

    if q and q.strip():
        params["q"] = f"%{q.strip()}%"
        search_cols = [
            "COALESCE(cc.serial, '') ILIKE :q",
            "COALESCE(cc.system_name, '') ILIKE :q",
            "COALESCE(cc.a_patchpanel_id, '') ILIKE :q",
            "COALESCE(cc.a_port_label, '') ILIKE :q",
            "COALESCE(CAST(cc.customer_patchpanel_id AS TEXT), '') ILIKE :q",
            "COALESCE(cc.customer_port_label, '') ILIKE :q",
            "COALESCE(cc.z_pp_number, '') ILIKE :q",
            "COALESCE(cc.rack_code, '') ILIKE :q",
            "COALESCE(z_pp.instance_id, '') ILIKE :q",
            "COALESCE(cust.name, '') ILIKE :q",
        ]
        if _has_column(db, "cross_connects", "serial_number"):
            search_cols.insert(1, "COALESCE(cc.serial_number, '') ILIKE :q")
        if _has_column(db, "cross_connects", "switch_name"):
            search_cols.append("COALESCE(cc.switch_name, '') ILIKE :q")
        if _has_column(db, "cross_connects", "switch_port"):
            search_cols.append("COALESCE(cc.switch_port, '') ILIKE :q")
        if _has_column(db, "cross_connects", "customer_name"):
            search_cols.append("COALESCE(cc.customer_name, '') ILIKE :q")
        where.append(
            "(" + " OR ".join(search_cols) + ")"
        )

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = _line_select_sql(db) + f" {where_sql} ORDER BY cc.created_at DESC, cc.id DESC LIMIT :limit"
    rows = db.execute(text(sql), params).mappings().all()

    items: list[dict[str, Any]] = []
    for row in rows:
        line = dict(row)
        line["serial"] = line.get("serial_effective") or line.get("serial")
        line.pop("serial_effective", None)
        line = _swap_backbone_fields(line) or line
        items.append(_line_to_public(line))

    return {"success": True, "total": len(items), "items": items}


@router.get("/cross_connects/archive")
def list_archived_cross_connects(
    q: str | None = Query(None, description="Search"),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """List deinstalled cross-connects from the archive table (documentation only)."""
    _ensure_kw_flow_tables(db)
    params: dict[str, Any] = {"limit": int(limit)}
    where = []
    if q and q.strip():
        params["q"] = f"%{q.strip()}%"
        where.append(
            "(COALESCE(a.serial, '') ILIKE :q"
            " OR COALESCE(a.system_name, '') ILIKE :q"
            " OR COALESCE(a.switch_name, '') ILIKE :q"
            " OR COALESCE(a.a_patchpanel_id, '') ILIKE :q"
            " OR COALESCE(CAST(a.customer_patchpanel_id AS TEXT), '') ILIKE :q"
            " OR COALESCE(a.reason, '') ILIKE :q)"
        )
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""
        SELECT a.*,
               z_pp.instance_id AS customer_patchpanel_instance_id,
               cust.name AS customer_base_name
        FROM public.cross_connects_archive a
        LEFT JOIN public.patchpanel_instances z_pp
               ON z_pp.id = a.customer_patchpanel_id
        LEFT JOIN public.customers cust
               ON cust.id = z_pp.customer_id
        {where_sql}
        ORDER BY a.deinstalled_at DESC, a.id DESC
        LIMIT :limit
    """
    rows = db.execute(text(sql), params).mappings().all()
    items = []
    for row in rows:
        r = dict(row)
        z_pp = r.get("customer_patchpanel_instance_id") or r.get("customer_patchpanel_id")
        sys_name = r.get("system_name") or ""
        base_name = r.get("customer_base_name") or ""
        customer = sys_name if (":" in sys_name) else (sys_name or base_name or "-")
        items.append({
            "id": int(r["id"]),
            "original_id": r.get("original_id"),
            "serial": r.get("serial") or r.get("serial_number"),
            "product_id": r.get("product_id"),
            "customer": customer,
            "system_name": r.get("system_name"),
            "switch_name": r.get("switch_name"),
            "switch_port": r.get("switch_port"),
            "a_side": {
                "pp": r.get("a_patchpanel_id"),
                "port": r.get("a_port_label"),
            },
            "z_side": {
                "pp": z_pp,
                "port": r.get("customer_port_label"),
            },
            "bb_in": {
                "pp": r.get("backbone_in_instance_id"),
                "port": r.get("backbone_in_port_label"),
            },
            "bb_out": {
                "pp": r.get("backbone_out_instance_id"),
                "port": r.get("backbone_out_port_label"),
            },
            "status": r.get("status"),
            "deinstalled_at": str(r["deinstalled_at"]) if r.get("deinstalled_at") else None,
            "deinstalled_by": r.get("deinstalled_by"),
            "reason": r.get("reason"),
            "original_created_at": str(r["original_created_at"]) if r.get("original_created_at") else None,
        })
    return {"success": True, "total": len(items), "items": items}


@router.get("/kw_plans")
def list_kw_plans_minimal(
    status: str | None = Query(None, description="open|locked|completed or CSV"),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)

    statuses = []
    if status and status.strip():
        statuses = [_normalize_plan_status(x) for x in status.split(",") if x.strip()]
        statuses = [x for x in statuses if x in PLAN_STATUS_ALLOWED]

    params: dict[str, Any] = {"limit": int(limit)}
    where_sql = ""
    if statuses:
        params["statuses"] = statuses
        where_sql = "WHERE p.status = ANY(:statuses)"

    rows = db.execute(
        text(
            f"""
            SELECT
                p.id,
                p.year,
                p.kw,
                p.status,
                p.created_by,
                p.created_at,
                COUNT(c.id) AS changes_total,
                COUNT(c.id) FILTER (WHERE c.status IN ('planned', 'in_progress')) AS open_changes
            FROM public.kw_plans p
            LEFT JOIN public.kw_changes c ON c.kw_plan_id = p.id
            {where_sql}
            GROUP BY p.id, p.year, p.kw, p.status, p.created_by, p.created_at
            ORDER BY p.year DESC, p.kw DESC
            LIMIT :limit
            """
        ),
        params,
    ).mappings().all()

    return {"success": True, "items": [_plan_row_to_out(dict(row)) for row in rows]}


@router.post("/kw_plans")
def create_kw_plan_minimal(
    payload: KwPlanCreateIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)
    year, kw = _parse_kw_label(payload.kw)
    status = _normalize_plan_status(payload.status)
    user_id = _get_user_id(current_user)

    existing = db.execute(
        text(
            """
            SELECT id, year, kw, status, created_by, created_at
            FROM public.kw_plans
            WHERE year = :year AND kw = :kw
            LIMIT 1
            """
        ),
        {"year": year, "kw": kw},
    ).mappings().first()

    if existing:
        row = dict(existing)
        if _normalize_plan_status(row.get("status")) != status and status in PLAN_STATUS_ALLOWED:
            db.execute(
                text("UPDATE public.kw_plans SET status = :status WHERE id = :id"),
                {"status": status, "id": int(row["id"])},
            )
            row["status"] = status
        row["changes_total"] = 0
        row["open_changes"] = 0
        return {"success": True, "created": False, "plan": _plan_row_to_out(row)}

    inserted = db.execute(
        text(
            """
            INSERT INTO public.kw_plans (year, kw, status, created_by)
            VALUES (:year, :kw, :status, :created_by)
            RETURNING id, year, kw, status, created_by, created_at
            """
        ),
        {
            "year": year,
            "kw": kw,
            "status": status,
            "created_by": user_id,
        },
    ).mappings().first()

    row = dict(inserted)
    row["changes_total"] = 0
    row["open_changes"] = 0
    return {"success": True, "created": True, "plan": _plan_row_to_out(row)}


@router.patch("/kw_plans/{plan_id}")
def update_kw_plan_minimal(
    payload: KwPlanPatchIn,
    plan_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)
    status = _normalize_plan_status(payload.status)

    row = db.execute(
        text(
            """
            UPDATE public.kw_plans
            SET status = :status
            WHERE id = :id
            RETURNING id, year, kw, status, created_by, created_at
            """
        ),
        {"id": int(plan_id), "status": status},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="KW plan not found")
    out = dict(row)
    out["changes_total"] = 0
    out["open_changes"] = 0
    return {"success": True, "plan": _plan_row_to_out(out)}


# ────────────────────────────────────────────────────
# KW Plan: Complete (finish) and Excel report
# ────────────────────────────────────────────────────

@router.post("/kw_plans/{plan_id}/complete")
def complete_kw_plan(
    plan_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Mark a KW plan as completed.
    Once completed, no more changes can be created or applied.
    All non-done changes are automatically set to 'canceled'.
    """
    _ensure_kw_flow_tables(db)

    plan = db.execute(
        text("SELECT id, year, kw, status, created_by, created_at FROM public.kw_plans WHERE id = :id FOR UPDATE"),
        {"id": int(plan_id)},
    ).mappings().first()
    if not plan:
        raise HTTPException(status_code=404, detail="KW plan not found")
    plan = dict(plan)

    if _normalize_plan_status(plan.get("status")) == "completed":
        raise HTTPException(status_code=409, detail="KW plan is already completed")

    # Check: all changes should be done or canceled
    open_changes = db.execute(
        text("SELECT COUNT(*) FROM public.kw_changes WHERE kw_plan_id = :pid AND LOWER(status) NOT IN ('done','canceled')"),
        {"pid": int(plan_id)},
    ).scalar() or 0

    if open_changes > 0:
        # Auto-cancel remaining open changes
        db.execute(
            text("""UPDATE public.kw_changes
                    SET status = 'canceled', completed_at = NOW()
                    WHERE kw_plan_id = :pid AND LOWER(status) NOT IN ('done','canceled')"""),
            {"pid": int(plan_id)},
        )

    upd_sql = "UPDATE public.kw_plans SET status = 'completed'"
    if _has_column(db, "kw_plans", "completed_at"):
        upd_sql += ", completed_at = NOW()"
    upd_sql += " WHERE id = :id RETURNING id, year, kw, status, created_by, created_at"

    row = db.execute(text(upd_sql), {"id": int(plan_id)}).mappings().first()
    out = dict(row)
    out["changes_total"] = 0
    out["open_changes"] = 0
    return {
        "success": True,
        "plan": _plan_row_to_out(out),
        "canceled_open": int(open_changes),
    }


@router.get("/kw_plans/{plan_id}/report.xlsx")
def download_kw_report(
    plan_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Download an Excel report for a completed KW plan.
    The report has four sheets: Installation, Deinstallation, Line Move, Path Move.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _ensure_kw_flow_tables(db)

    plan = db.execute(
        text("SELECT id, year, kw, status, created_by, created_at FROM public.kw_plans WHERE id = :id"),
        {"id": int(plan_id)},
    ).mappings().first()
    if not plan:
        raise HTTPException(status_code=404, detail="KW plan not found")
    plan = dict(plan)

    if _normalize_plan_status(plan.get("status")) != "completed":
        raise HTTPException(status_code=403, detail="Report is only available for completed KW plans")

    kw_label = _kw_label(int(plan["year"]), int(plan["kw"]))

    changes = db.execute(
        text("""
            SELECT c.id, c.type, c.target_cross_connect_id, c.payload_json,
                   c.status, c.created_at, c.completed_at, c.created_by, c.completed_by
            FROM public.kw_changes c
            WHERE c.kw_plan_id = :pid
            ORDER BY c.type, c.id
        """),
        {"pid": int(plan_id)},
    ).mappings().all()
    changes = [dict(r) for r in changes]

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    done_fill = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")
    canceled_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")

    def _safe(v):
        if isinstance(v, datetime):
            return v.replace(tzinfo=None) if v.tzinfo else v
        return v

    def _style_sheet(ws, headers, rows):
        """Apply headers, rows, and formatting to a worksheet."""
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_data in rows:
            ws.append([_safe(v) for v in row_data])
        for ri in range(2, ws.max_row + 1):
            status_col = None
            for ci, h in enumerate(headers, 1):
                if h == "Status":
                    status_col = ci
                    break
            for cell in ws[ri]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            if status_col:
                s_val = str(ws.cell(row=ri, column=status_col).value or "").lower()
                fill = done_fill if s_val == "done" else canceled_fill if s_val == "canceled" else None
                if fill:
                    for cell in ws[ri]:
                        cell.fill = fill
        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # ── Sheet 1: Installation ──
    ws_inst = wb.active
    ws_inst.title = "Installation"
    inst_headers = ["ID", "Serial", "Switch Name", "Switch Port",
                    "A Patchpanel", "A Port",
                    "BB IN PP", "BB IN Port", "BB OUT PP", "BB OUT Port",
                    "Z Patchpanel", "Z Port", "Kunde", "Status",
                    "Erstellt", "Abgeschlossen"]
    inst_rows = []
    for c in changes:
        if str(c.get("type") or "").upper() != "NEW_INSTALL":
            continue
        p = c.get("payload_json") or {}
        l = p.get("new_line") or p
        inst_rows.append([
            c["id"],
            l.get("serial") or l.get("product_id") or "-",
            l.get("switch_name") or "-",
            l.get("switch_port") or "-",
            l.get("a_patchpanel_id") or "-",
            l.get("a_port_label") or "-",
            l.get("backbone_in_instance_id") or "-",
            l.get("backbone_in_port_label") or "-",
            l.get("backbone_out_instance_id") or "-",
            l.get("backbone_out_port_label") or "-",
            l.get("customer_patchpanel_instance_id") or l.get("customer_patchpanel_id") or "-",
            l.get("customer_port_label") or "-",
            l.get("system_name") or "-",
            str(c.get("status") or "").lower(),
            c.get("created_at"),
            c.get("completed_at"),
        ])
    _style_sheet(ws_inst, inst_headers, inst_rows)

    # ── Sheet 2: Deinstallation ──
    ws_deinst = wb.create_sheet("Deinstallation")
    deinst_headers = ["ID", "Leitung ID", "Serial (vorher)", "Kunde (vorher)", "Grund", "Status",
                      "Erstellt", "Abgeschlossen"]
    deinst_rows = []
    for c in changes:
        if str(c.get("type") or "").upper() != "DEINSTALL":
            continue
        p = c.get("payload_json") or {}
        snap = p.get("snapshot_before") or {}
        deinst_rows.append([
            c["id"],
            c.get("target_cross_connect_id") or "-",
            snap.get("serial") or "-",
            snap.get("system_name") or "-",
            p.get("reason") or "-",
            str(c.get("status") or "").lower(),
            c.get("created_at"),
            c.get("completed_at"),
        ])
    _style_sheet(ws_deinst, deinst_headers, deinst_rows)

    # ── Sheet 3: Line Move ──
    ws_lm = wb.create_sheet("Line Move")
    lm_headers = ["ID", "Serial", "Leitung ID", "Alter Z PP", "Alter Z Port",
                   "Neuer Z PP", "Neuer Z Port",
                   "Alter BB IN", "Alter BB OUT",
                   "Neuer BB IN", "Neuer BB OUT",
                   "Rack", "Status", "Erstellt", "Abgeschlossen"]
    lm_rows = []
    for c in changes:
        if str(c.get("type") or "").upper() != "LINE_MOVE":
            continue
        p = c.get("payload_json") or {}
        oz = p.get("old_z") or {}
        nz = p.get("new_z") or {}
        ob = p.get("old_bb") or {}
        snap = p.get("snapshot") or {}
        lm_rows.append([
            c["id"],
            snap.get("serial") or "-",
            c.get("target_cross_connect_id") or "-",
            oz.get("customer_patchpanel_instance_id") or oz.get("customer_patchpanel_id") or "-",
            oz.get("customer_port_label") or "-",
            nz.get("customer_patchpanel_instance_id") or nz.get("customer_patchpanel_id") or "-",
            nz.get("customer_port_label") or "-",
            f'{ob.get("backbone_in_instance_id", "-")}/{ob.get("backbone_in_port_label", "-")}',
            f'{ob.get("backbone_out_instance_id", "-")}/{ob.get("backbone_out_port_label", "-")}',
            f'{nz.get("backbone_in_instance_id", "-")}/{nz.get("backbone_in_port_label", "-")}' if nz.get("backbone_in_instance_id") else "-",
            f'{nz.get("backbone_out_instance_id", "-")}/{nz.get("backbone_out_port_label", "-")}' if nz.get("backbone_out_instance_id") else "-",
            nz.get("rack_code") or "-",
            str(c.get("status") or "").lower(),
            c.get("created_at"),
            c.get("completed_at"),
        ])
    _style_sheet(ws_lm, lm_headers, lm_rows)

    # ── Sheet 4: Path Move ──
    ws_pm = wb.create_sheet("Path Move")
    pm_headers = ["ID", "Leitung A Serial", "Leitung B Serial",
                   "A BB IN (alt)", "A BB OUT (alt)",
                   "B BB IN (alt)", "B BB OUT (alt)",
                   "Status", "Erstellt", "Abgeschlossen"]
    pm_rows = []
    for c in changes:
        if str(c.get("type") or "").upper() != "PATH_MOVE":
            continue
        p = c.get("payload_json") or {}
        a_bb = p.get("line_a_old_bb") or {}
        b_bb = p.get("line_b_old_bb") or {}
        pm_rows.append([
            c["id"],
            p.get("line_a_serial") or str(p.get("line_a_id") or "-"),
            p.get("line_b_serial") or str(p.get("line_b_id") or "-"),
            f'{a_bb.get("backbone_in_instance_id", "-")}/{a_bb.get("backbone_in_port_label", "-")}',
            f'{a_bb.get("backbone_out_instance_id", "-")}/{a_bb.get("backbone_out_port_label", "-")}',
            f'{b_bb.get("backbone_in_instance_id", "-")}/{b_bb.get("backbone_in_port_label", "-")}',
            f'{b_bb.get("backbone_out_instance_id", "-")}/{b_bb.get("backbone_out_port_label", "-")}',
            str(c.get("status") or "").lower(),
            c.get("created_at"),
            c.get("completed_at"),
        ])
    _style_sheet(ws_pm, pm_headers, pm_rows)

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Zusammenfassung", 0)
    ws_sum.append(["KW Report"])
    ws_sum["A1"].font = Font(bold=True, size=16)
    ws_sum.append([])
    ws_sum.append(["Kalenderwoche", kw_label])
    ws_sum.append(["Status", "Abgeschlossen"])
    ws_sum.append(["Abgeschlossen am", _safe(plan.get("completed_at") or datetime.now())])
    ws_sum.append([])
    ws_sum.append(["Typ", "Gesamt", "Erledigt", "Abgebrochen"])
    type_order = ["NEW_INSTALL", "DEINSTALL", "LINE_MOVE", "PATH_MOVE"]
    type_labels = {"NEW_INSTALL": "Installation", "DEINSTALL": "Deinstallation", "LINE_MOVE": "Line Move", "PATH_MOVE": "Path Move"}
    for t in type_order:
        typed = [c for c in changes if str(c.get("type") or "").upper() == t]
        done_cnt = sum(1 for c in typed if str(c.get("status") or "").lower() == "done")
        canceled_cnt = sum(1 for c in typed if str(c.get("status") or "").lower() == "canceled")
        ws_sum.append([type_labels.get(t, t), len(typed), done_cnt, canceled_cnt])
    ws_sum.append([])
    ws_sum.append(["Gesamt", len(changes),
                   sum(1 for c in changes if str(c.get("status") or "").lower() == "done"),
                   sum(1 for c in changes if str(c.get("status") or "").lower() == "canceled")])
    for row in ws_sum.iter_rows(min_row=7, max_row=7, max_col=4):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
    for ri in range(3, ws_sum.max_row + 1):
        for cell in ws_sum[ri]:
            cell.border = thin_border
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 14

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"KW_Report_{kw_label}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/kw_changes")
def list_kw_changes_minimal(
    kw: str | None = Query(None),
    kw_plan_id: int | None = Query(None, ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)

    if not kw and not kw_plan_id:
        raise HTTPException(status_code=400, detail="kw or kw_plan_id is required")

    params: dict[str, Any] = {}
    where = []
    if kw_plan_id:
        where.append("c.kw_plan_id = :kw_plan_id")
        params["kw_plan_id"] = int(kw_plan_id)
    if kw:
        year, week = _parse_kw_label(kw)
        where.append("p.year = :year AND p.kw = :kw")
        params["year"] = year
        params["kw"] = week

    where_sql = "WHERE " + " AND ".join(where)
    rows = db.execute(
        text(
            f"""
            SELECT
                c.id,
                c.kw_plan_id,
                c.type,
                c.target_cross_connect_id,
                c.payload_json,
                c.status,
                c.created_at,
                c.completed_at,
                c.created_by,
                c.completed_by,
                p.year,
                p.kw
            FROM public.kw_changes c
            JOIN public.kw_plans p ON p.id = c.kw_plan_id
            {where_sql}
            ORDER BY c.created_at DESC, c.id DESC
            """
        ),
        params,
    ).mappings().all()

    return {"success": True, "items": [_change_row_to_out(dict(row)) for row in rows]}


@router.post("/kw_changes")
def create_kw_change_minimal(
    payload: KwChangeCreateIn,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)
    user_id = _get_user_id(current_user)

    change_type = str(payload.type or "").strip().upper()
    if change_type not in CHANGE_TYPE_ALLOWED:
        raise HTTPException(status_code=400, detail="Invalid change type")

    status = _validate_change_status(payload.status)
    if status == "done":
        raise HTTPException(status_code=400, detail="Use apply endpoint to complete a change")

    plan: dict[str, Any] | None = None
    if payload.kw_plan_id:
        plan = _find_plan_by_id(db, int(payload.kw_plan_id), for_update=False)
    elif payload.kw:
        plan = _find_plan_by_kw(db, payload.kw, for_update=False)

    if not plan and payload.kw:
        year, week = _parse_kw_label(payload.kw)
        row = db.execute(
            text(
                """
                INSERT INTO public.kw_plans (year, kw, status, created_by)
                VALUES (:year, :kw, 'open', :created_by)
                ON CONFLICT (year, kw) DO UPDATE SET year = EXCLUDED.year
                RETURNING id, year, kw, status, created_by, created_at
                """
            ),
            {"year": year, "kw": week, "created_by": user_id},
        ).mappings().first()
        plan = dict(row)

    if not plan:
        raise HTTPException(status_code=404, detail="KW plan not found")

    _assert_plan_writable(plan)

    payload_json = dict(payload.payload_json or {})
    target_id = int(payload.target_cross_connect_id) if payload.target_cross_connect_id else None

    if change_type == "NEW_INSTALL":
        line = payload_json.get("new_line") or payload_json
        serial = str(line.get("serial") or "").strip()
        product_id = str(line.get("product_id") or "").strip()
        if not serial and not product_id:
            raise HTTPException(status_code=400, detail="NEW_INSTALL requires serial or product_id in payload_json.new_line")
        payload_json["new_line"] = line
        target_id = None

    elif change_type == "DEINSTALL":
        if not target_id:
            raise HTTPException(status_code=400, detail="DEINSTALL requires target_cross_connect_id")
        line = _assert_line_active(_line_by_id(db, target_id, for_update=False), "Target line")
        payload_json.setdefault("snapshot_before", line)

    elif change_type == "LINE_MOVE":
        if not target_id:
            raise HTTPException(status_code=400, detail="LINE_MOVE requires target_cross_connect_id")
        line = _assert_line_active(_line_by_id(db, target_id, for_update=False), "Target line")

        new_z = payload_json.get("new_z") or {}
        pp = new_z.get("customer_patchpanel_id")
        port = str(new_z.get("customer_port_label") or "").strip()
        if not pp or not port:
            raise HTTPException(status_code=400, detail="LINE_MOVE requires payload_json.new_z customer_patchpanel_id + customer_port_label")

        _assert_z_port_free(db, int(pp), port, exclude_line_id=target_id)

        # Snapshot current line state for detail view
        # Resolve customer name (same logic as _line_to_public)
        _sn = line.get("system_name") or ""
        _rsn = line.get("resolved_system_name") or ""
        _bn = line.get("customer_base_name") or ""
        if ":" in _sn:
            _customer = _sn
        elif ":" in _rsn:
            _customer = _rsn
        else:
            _customer = _sn or _bn or "-"
        payload_json["snapshot"] = {
            "serial": line.get("serial") or line.get("serial_number"),
            "system_name": line.get("system_name"),
            "customer": _customer,
            "switch_name": line.get("switch_name"),
            "switch_port": line.get("switch_port"),
            "a_patchpanel_id": line.get("a_patchpanel_id"),
            "a_port_label": line.get("a_port_label"),
        }
        payload_json["old_z"] = payload_json.get("old_z") or {
            "customer_patchpanel_id": line.get("customer_patchpanel_id"),
            "customer_patchpanel_instance_id": line.get("customer_patchpanel_instance_id"),
            "customer_port_label": line.get("customer_port_label"),
            "z_pp_number": line.get("z_pp_number"),
            "rack_code": line.get("rack_code"),
        }
        payload_json["old_bb"] = {
            "backbone_in_instance_id": line.get("backbone_in_instance_id"),
            "backbone_in_port_label": line.get("backbone_in_port_label"),
            "backbone_out_instance_id": line.get("backbone_out_instance_id"),
            "backbone_out_port_label": line.get("backbone_out_port_label"),
        }
        payload_json["new_z"] = {
            "customer_patchpanel_id": int(pp),
            "customer_patchpanel_instance_id": new_z.get("customer_patchpanel_instance_id"),
            "customer_port_label": port,
            "z_pp_number": new_z.get("z_pp_number"),
            "rack_code": new_z.get("rack_code"),
            "backbone_in_instance_id": new_z.get("backbone_in_instance_id"),
            "backbone_in_port_label": new_z.get("backbone_in_port_label"),
            "backbone_out_instance_id": new_z.get("backbone_out_instance_id"),
            "backbone_out_port_label": new_z.get("backbone_out_port_label"),
        }

    elif change_type == "PATH_MOVE":
        line_a_id = payload_json.get("line_a_id") or target_id
        line_b_id = payload_json.get("line_b_id")
        if not line_a_id or not line_b_id:
            raise HTTPException(status_code=400, detail="PATH_MOVE requires payload_json.line_a_id + line_b_id")
        if int(line_a_id) == int(line_b_id):
            raise HTTPException(status_code=400, detail="PATH_MOVE requires two different active lines")

        line_a = _assert_line_active(_line_by_id(db, int(line_a_id), for_update=False), "line_a")
        line_b = _assert_line_active(_line_by_id(db, int(line_b_id), for_update=False), "line_b")
        target_id = int(line_a_id)
        payload_json["line_a_id"] = int(line_a_id)
        payload_json["line_b_id"] = int(line_b_id)
        # Persist serial numbers for display purposes
        payload_json.setdefault("line_a_serial", line_a.get("serial") or line_a.get("serial_number") or None)
        payload_json.setdefault("line_b_serial", line_b.get("serial") or line_b.get("serial_number") or None)
        payload_json["line_a_old_bb"] = payload_json.get("line_a_old_bb") or {
            "backbone_in_instance_id": line_a.get("backbone_in_instance_id"),
            "backbone_in_port_label": line_a.get("backbone_in_port_label"),
            "backbone_out_instance_id": line_a.get("backbone_out_instance_id"),
            "backbone_out_port_label": line_a.get("backbone_out_port_label"),
        }
        payload_json["line_b_old_bb"] = payload_json.get("line_b_old_bb") or {
            "backbone_in_instance_id": line_b.get("backbone_in_instance_id"),
            "backbone_in_port_label": line_b.get("backbone_in_port_label"),
            "backbone_out_instance_id": line_b.get("backbone_out_instance_id"),
            "backbone_out_port_label": line_b.get("backbone_out_port_label"),
        }

    row = db.execute(
        text(
            """
            INSERT INTO public.kw_changes (
                kw_plan_id,
                type,
                target_cross_connect_id,
                payload_json,
                status,
                created_by
            )
            VALUES (
                :kw_plan_id,
                :type,
                :target_cross_connect_id,
                CAST(:payload_json AS jsonb),
                :status,
                :created_by
            )
            RETURNING id, kw_plan_id, type, target_cross_connect_id, payload_json, status, created_at, completed_at, created_by, completed_by
            """
        ),
        {
            "kw_plan_id": int(plan["id"]),
            "type": change_type,
            "target_cross_connect_id": target_id,
            "payload_json": _serialize_json(payload_json),
            "status": status,
            "created_by": user_id,
        },
    ).mappings().first()

    out = dict(row)
    out["year"] = int(plan["year"])
    out["kw"] = int(plan["kw"])
    return {"success": True, "change": _change_row_to_out(out)}


@router.patch("/kw_changes/{change_id}")
def update_kw_change_minimal(
    payload: KwChangePatchIn,
    change_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)

    row = db.execute(
        text(
            """
            SELECT
                c.id,
                c.kw_plan_id,
                c.type,
                c.status,
                c.payload_json,
                p.year,
                p.kw,
                p.status AS plan_status
            FROM public.kw_changes c
            JOIN public.kw_plans p ON p.id = c.kw_plan_id
            WHERE c.id = :id
            LIMIT 1
            """
        ),
        {"id": int(change_id)},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="KW change not found")

    change = dict(row)
    if str(change.get("status") or "").lower() == "done":
        raise HTTPException(status_code=409, detail="Done change cannot be edited")

    if _normalize_plan_status(change.get("plan_status")) in PLAN_EDIT_BLOCKED:
        raise HTTPException(status_code=403, detail="KW plan is locked/completed")

    updates = []
    params: dict[str, Any] = {"id": int(change_id)}
    new_status = None
    if payload.status is not None:
        new_status = _validate_change_status(payload.status)
        updates.append("status = :status")
        params["status"] = new_status
    if payload.payload_json is not None:
        updates.append("payload_json = CAST(:payload_json AS jsonb)")
        params["payload_json"] = _serialize_json(payload.payload_json)
    if not updates:
        change["plan_status"] = _normalize_plan_status(change.get("plan_status"))
        return {"success": True, "change": _change_row_to_out(change)}

    if new_status in {"canceled", "done"}:
        updates.append("completed_at = NOW()")
    elif new_status in {"planned", "in_progress"}:
        updates.append("completed_at = NULL")

    updated = db.execute(
        text(
            "UPDATE public.kw_changes SET "
            + ", ".join(updates)
            + " WHERE id = :id "
            + "RETURNING id, kw_plan_id, type, target_cross_connect_id, payload_json, status, created_at, completed_at, created_by, completed_by"
        ),
        params,
    ).mappings().first()

    out = dict(updated)
    out["year"] = int(change["year"])
    out["kw"] = int(change["kw"])
    return {"success": True, "change": _change_row_to_out(out)}


@router.delete("/kw_changes/{change_id}")
def delete_kw_change(
    change_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Delete a kw_change that has not yet been applied (status != done)."""
    _ensure_kw_flow_tables(db)

    row = db.execute(
        text(
            """
            SELECT c.id, c.status, p.status AS plan_status
            FROM public.kw_changes c
            JOIN public.kw_plans p ON p.id = c.kw_plan_id
            WHERE c.id = :id
            """
        ),
        {"id": int(change_id)},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="KW change not found")

    if str(row.get("status") or "").lower() == "done":
        raise HTTPException(status_code=409, detail="Ausgefuehrte Massnahme kann nicht geloescht werden")

    if _normalize_plan_status(row.get("plan_status")) in PLAN_EDIT_BLOCKED:
        raise HTTPException(status_code=403, detail="KW plan is locked/completed")

    db.execute(
        text("DELETE FROM public.kw_changes WHERE id = :id"),
        {"id": int(change_id)},
    )
    db.commit()
    return {"success": True, "deleted": True}


@router.post("/kw_changes/{change_id}/apply")
def apply_kw_change_minimal(
    change_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ensure_kw_flow_tables(db)
    user_id = _get_user_id(current_user)

    row = db.execute(
        text(
            """
            SELECT
                c.id,
                c.kw_plan_id,
                c.type,
                c.target_cross_connect_id,
                c.payload_json,
                c.status,
                p.year,
                p.kw,
                p.status AS plan_status
            FROM public.kw_changes c
            JOIN public.kw_plans p ON p.id = c.kw_plan_id
            WHERE c.id = :id
            FOR UPDATE OF c
            """
        ),
        {"id": int(change_id)},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="KW change not found")

    change = dict(row)
    change_status = str(change.get("status") or "").lower().replace("cancelled", "canceled")
    if change_status not in CHANGE_APPLY_ALLOWED:
        raise HTTPException(status_code=409, detail="APPLY allowed only for planned/in_progress changes")

    plan_status = _normalize_plan_status(change.get("plan_status"))
    if plan_status == "completed":
        raise HTTPException(status_code=403, detail="KW plan completed; apply blocked")

    change_type = str(change.get("type") or "").strip().upper()
    payload_json = dict(change.get("payload_json") or {})
    target = change.get("target_cross_connect_id")
    created_line_id: int | None = None
    try:
        if change_type == "NEW_INSTALL":
            line = payload_json.get("new_line") or payload_json
            serial = str(line.get("serial") or "").strip()
            switch_name = str(line.get("switch_name") or "KW_PLAN").strip()
            switch_port = str(line.get("switch_port") or "KW_PLAN").strip()
            a_pp = str(line.get("a_patchpanel_id") or line.get("a_pp") or "").strip()
            a_port = str(line.get("a_port_label") or line.get("a_port") or "").strip()
            bb_in_pp = str(line.get("backbone_in_instance_id") or line.get("bb_in_pp") or "").strip()
            bb_in_port = str(line.get("backbone_in_port_label") or line.get("bb_in_port") or "").strip()
            bb_out_pp = str(line.get("backbone_out_instance_id") or line.get("bb_out_pp") or "").strip()
            bb_out_port = str(line.get("backbone_out_port_label") or line.get("bb_out_port") or "").strip()
            z_pp_id = line.get("customer_patchpanel_id") or line.get("z_pp")
            z_port = str(line.get("customer_port_label") or line.get("z_port") or "").strip()
            product_id = str(line.get("product_id") or "").strip()
            z_pp_instance_id = str(line.get("customer_patchpanel_instance_id") or line.get("z_pp_number") or "").strip()

            if not serial:
                serial = line.get("product_id") or line.get("sales_order") or ""
            if not serial:
                raise HTTPException(status_code=400, detail="NEW_INSTALL requires serial or product_id")
            # Use defaults for optional fields so the cross-connect can still be created
            if not a_pp:
                a_pp = line.get("a_patchpanel_id") or "KW_PLAN"
            if not a_port:
                a_port = line.get("a_port_label") or "-"
            if not z_pp_id or not z_port:
                raise HTTPException(status_code=400, detail="NEW_INSTALL requires Z-side (customer_patchpanel_id + customer_port_label)")
            try:
                z_pp_id_int = int(z_pp_id)
            except Exception as exc:
                raise HTTPException(status_code=400, detail="NEW_INSTALL requires numeric customer_patchpanel_id") from exc

            _assert_z_port_free(db, z_pp_id_int, z_port, exclude_line_id=None)
            if a_pp and a_pp != "KW_PLAN" and a_port and a_port != "-":
                _assert_field_pair_free(db, "a_patchpanel_id", "a_port_label", a_pp, a_port, None)
            if bb_out_pp and bb_out_port:
                _assert_field_pair_free(db, "backbone_in_instance_id", "backbone_in_port_label", bb_out_pp, bb_out_port, None)
            if bb_in_pp and bb_in_port:
                _assert_field_pair_free(db, "backbone_out_instance_id", "backbone_out_port_label", bb_in_pp, bb_in_port, None)

            line_for_db = _swap_backbone_payload(
                {
                    "backbone_in_instance_id": bb_in_pp,
                    "backbone_in_port_label": bb_in_port,
                    "backbone_out_instance_id": bb_out_pp,
                    "backbone_out_port_label": bb_out_port,
                }
            ) or {}

            # Resolve z_pp instance_id and rack_code from DB if not provided
            z_rack_code = ""
            if z_pp_id:
                pp_row = db.execute(text("SELECT instance_id, rack_label FROM patchpanel_instances WHERE id = :id"), {"id": int(z_pp_id)}).mappings().first()
                if pp_row:
                    if not z_pp_instance_id:
                        z_pp_instance_id = pp_row["instance_id"] or ""
                    z_rack_code = pp_row["rack_label"] or ""

            has_serial_number = _has_column(db, "cross_connects", "serial_number")

            columns = [
                "serial",
                "switch_name",
                "switch_port",
                "a_patchpanel_id",
                "a_port_label",
                "backbone_out_instance_id",
                "backbone_out_port_label",
                "backbone_in_instance_id",
                "backbone_in_port_label",
                "customer_patchpanel_id",
                "customer_port_label",
                "status",
            ]
            values = {
                "serial": serial,
                "switch_name": switch_name,
                "switch_port": switch_port,
                "a_patchpanel_id": a_pp,
                "a_port_label": a_port,
                "backbone_out_instance_id": line_for_db.get("backbone_out_instance_id") or "",
                "backbone_out_port_label": line_for_db.get("backbone_out_port_label") or "",
                "backbone_in_instance_id": line_for_db.get("backbone_in_instance_id") or "",
                "backbone_in_port_label": line_for_db.get("backbone_in_port_label") or "",
                "customer_patchpanel_id": z_pp_id_int,
                "customer_port_label": z_port,
                "status": "active",
            }

            if has_serial_number:
                columns.append("serial_number")
                values["serial_number"] = serial
            if _has_column(db, "cross_connects", "product_id"):
                columns.append("product_id")
                values["product_id"] = product_id or None
            if _has_column(db, "cross_connects", "system_name"):
                columns.append("system_name")
                values["system_name"] = str(line.get("system_name") or line.get("customer") or "").strip() or None
            if _has_column(db, "cross_connects", "rack_code"):
                columns.append("rack_code")
                values["rack_code"] = z_rack_code or str(line.get("rack_code") or line.get("z_rack") or "").strip() or None
            if _has_column(db, "cross_connects", "z_pp_number"):
                columns.append("z_pp_number")
                values["z_pp_number"] = z_pp_instance_id or None
            if _has_column(db, "cross_connects", "updated_at"):
                columns.append("updated_at")
                values["updated_at"] = datetime.utcnow()

            sql_cols = ", ".join(columns)
            sql_vals = ", ".join([f":{c}" for c in columns])
            created = db.execute(
                text(
                    f"""
                    INSERT INTO public.cross_connects ({sql_cols})
                    VALUES ({sql_vals})
                    RETURNING id
                    """
                ),
                values,
            ).first()
            created_line_id = int(created[0]) if created else None

        elif change_type == "DEINSTALL":
            if not target:
                raise HTTPException(status_code=400, detail="DEINSTALL target_cross_connect_id missing")
            line = _assert_line_active(_line_by_id(db, int(target), for_update=True), "Target line")
            reason = (payload_json.get("reason") or "") if payload_json else ""
            deinstalled_by = str(current_user.get("username", "")) if isinstance(current_user, dict) else getattr(current_user, "username", "")
            # Archive the line with ALL path info preserved as documentation
            db.execute(
                text(
                    """
                    INSERT INTO public.cross_connects_archive (
                        original_id, serial, serial_number, product_id,
                        switch_name, switch_port,
                        a_patchpanel_id, a_port_label,
                        backbone_out_instance_id, backbone_out_port_label,
                        backbone_in_instance_id, backbone_in_port_label,
                        customer_patchpanel_id, customer_port_label,
                        z_pp_number, rack_code, system_name,
                        customer_id, customer_rack_id, customer_location_id,
                        job_id, source_audit_line_id,
                        status, original_created_at,
                        deinstalled_at, deinstalled_by, reason
                    )
                    SELECT
                        id, serial, serial_number, product_id,
                        switch_name, switch_port,
                        a_patchpanel_id, a_port_label,
                        backbone_out_instance_id, backbone_out_port_label,
                        backbone_in_instance_id, backbone_in_port_label,
                        customer_patchpanel_id, customer_port_label,
                        z_pp_number, rack_code, system_name,
                        customer_id, customer_rack_id, customer_location_id,
                        job_id, source_audit_line_id,
                        'deinstalled', created_at,
                        NOW(), :deinstalled_by, :reason
                    FROM public.cross_connects
                    WHERE id = :id
                    """
                ),
                {"id": int(line["id"]), "deinstalled_by": deinstalled_by, "reason": reason},
            )
            # Delete from active cross_connects (ports are now free)
            db.execute(
                text("DELETE FROM public.cross_connects WHERE id = :id"),
                {"id": int(line["id"])},
            )

        elif change_type == "LINE_MOVE":
            if not target:
                raise HTTPException(status_code=400, detail="LINE_MOVE target_cross_connect_id missing")
            line = _assert_line_active(_line_by_id(db, int(target), for_update=True), "Target line")
            new_z = payload_json.get("new_z") or {}
            pp_id = new_z.get("customer_patchpanel_id")
            port = str(new_z.get("customer_port_label") or "").strip()
            if not pp_id or not port:
                raise HTTPException(status_code=400, detail="LINE_MOVE payload missing new_z")
            _assert_z_port_free(db, int(pp_id), port, exclude_line_id=int(line["id"]))

            # Update Z-side + set new BB IN/OUT (or clear if not provided)
            bb_in_inst = str(new_z.get("backbone_in_instance_id") or "").strip()
            bb_in_port = str(new_z.get("backbone_in_port_label") or "").strip()
            bb_out_inst = str(new_z.get("backbone_out_instance_id") or "").strip()
            bb_out_port = str(new_z.get("backbone_out_port_label") or "").strip()

            updates = [
                "customer_patchpanel_id = :pp_id",
                "customer_port_label = :port",
                "backbone_in_instance_id = :bb_in_inst",
                "backbone_in_port_label = :bb_in_port",
                "backbone_out_instance_id = :bb_out_inst",
                "backbone_out_port_label = :bb_out_port",
            ]
            params: dict[str, Any] = {
                "id": int(line["id"]),
                "pp_id": int(pp_id),
                "port": port,
                "bb_in_inst": bb_in_inst,
                "bb_in_port": bb_in_port,
                "bb_out_inst": bb_out_inst,
                "bb_out_port": bb_out_port,
            }
            # Also update customer_patchpanel_instance_id if column exists
            if _has_column(db, "cross_connects", "customer_patchpanel_instance_id"):
                updates.append("customer_patchpanel_instance_id = :cpp_inst")
                params["cpp_inst"] = str(new_z.get("customer_patchpanel_instance_id") or "").strip()
            if _has_column(db, "cross_connects", "z_pp_number"):
                updates.append("z_pp_number = :z_pp_number")
                params["z_pp_number"] = new_z.get("z_pp_number")
            if _has_column(db, "cross_connects", "rack_code"):
                updates.append("rack_code = :rack_code")
                params["rack_code"] = new_z.get("rack_code")
            if _has_column(db, "cross_connects", "updated_at"):
                updates.append("updated_at = NOW()")

            db.execute(
                text("UPDATE public.cross_connects SET " + ", ".join(updates) + " WHERE id = :id"),
                params,
            )

        elif change_type == "PATH_MOVE":
            line_a_id = payload_json.get("line_a_id") or target
            line_b_id = payload_json.get("line_b_id")
            if not line_a_id or not line_b_id:
                raise HTTPException(status_code=400, detail="PATH_MOVE requires line_a_id + line_b_id")
            if int(line_a_id) == int(line_b_id):
                raise HTTPException(status_code=400, detail="PATH_MOVE requires two different lines")

            line_a = _assert_line_active(_line_by_id(db, int(line_a_id), for_update=True), "line_a")
            line_b = _assert_line_active(_line_by_id(db, int(line_b_id), for_update=True), "line_b")

            a_old = {
                "backbone_in_instance_id": line_a.get("backbone_in_instance_id"),
                "backbone_in_port_label": line_a.get("backbone_in_port_label"),
                "backbone_out_instance_id": line_a.get("backbone_out_instance_id"),
                "backbone_out_port_label": line_a.get("backbone_out_port_label"),
            }
            b_old = {
                "backbone_in_instance_id": line_b.get("backbone_in_instance_id"),
                "backbone_in_port_label": line_b.get("backbone_in_port_label"),
                "backbone_out_instance_id": line_b.get("backbone_out_instance_id"),
                "backbone_out_port_label": line_b.get("backbone_out_port_label"),
            }

            a_new_db = _swap_backbone_payload(dict(b_old)) or {}
            b_new_db = _swap_backbone_payload(dict(a_old)) or {}

            has_upd = _has_column(db, "cross_connects", "updated_at")
            upd_frag = ", updated_at = NOW()" if has_upd else ""
            id_a = int(line_a["id"])
            id_b = int(line_b["id"])

            # --- Three-step swap to avoid partial unique index conflicts ---
            # Step 1: Remove line A from the partial unique index (status != 'active')
            db.execute(
                text("UPDATE public.cross_connects SET status = 'in_progress' WHERE id = :id"),
                {"id": id_a},
            )
            # Step 2: Update line B with A's old backbone values (no conflict, A is not active)
            db.execute(
                text(f"""UPDATE public.cross_connects
                         SET backbone_in_instance_id  = :bi_i,
                             backbone_in_port_label   = :bi_p,
                             backbone_out_instance_id  = :bo_i,
                             backbone_out_port_label   = :bo_p
                             {upd_frag}
                         WHERE id = :id"""),
                {"id": id_b,
                 "bi_i": b_new_db.get("backbone_in_instance_id"),
                 "bi_p": b_new_db.get("backbone_in_port_label"),
                 "bo_i": b_new_db.get("backbone_out_instance_id"),
                 "bo_p": b_new_db.get("backbone_out_port_label")},
            )
            # Step 3: Update line A with B's old backbone values and restore status
            db.execute(
                text(f"""UPDATE public.cross_connects
                         SET backbone_in_instance_id  = :bi_i,
                             backbone_in_port_label   = :bi_p,
                             backbone_out_instance_id  = :bo_i,
                             backbone_out_port_label   = :bo_p,
                             status = 'active'
                             {upd_frag}
                         WHERE id = :id"""),
                {"id": id_a,
                 "bi_i": a_new_db.get("backbone_in_instance_id"),
                 "bi_p": a_new_db.get("backbone_in_port_label"),
                 "bo_i": a_new_db.get("backbone_out_instance_id"),
                 "bo_p": a_new_db.get("backbone_out_port_label")},
            )

        else:
            raise HTTPException(status_code=400, detail="Unsupported change type")

        db.execute(
            text(
                """
                UPDATE public.kw_changes
                SET status = 'done',
                    completed_at = NOW(),
                    completed_by = :completed_by
                WHERE id = :id
                """
            ),
            {"id": int(change_id), "completed_by": user_id},
        )
    except HTTPException:
        raise
    except Exception as exc:
        msg = str(exc).lower()
        if "duplicate key" in msg or "unique" in msg:
            raise HTTPException(status_code=409, detail="Port occupancy conflict while applying change") from exc
        raise HTTPException(status_code=500, detail=f"Apply failed: {str(exc)}") from exc

    return {
        "success": True,
        "change_id": int(change_id),
        "status": "done",
        "created_cross_connect_id": created_line_id,
    }
